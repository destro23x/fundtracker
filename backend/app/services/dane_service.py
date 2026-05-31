"""
Serwis do parsowania pliku portfelowego i generowania znormalizowanego xlsx.

Znormalizowany format (kolumny xlsx / DB):
  Nazwa funduszu
  Nazwa subfunduszu
  Identyfikator funduszu lub subfunduszu
  Typ Funduszu
  Emitent
  Kraj emitenta
  Kod ISIN instrumentu
  Typ instrumentu
  Ilość instrumentów w portfelu
  Waluta wyceny instrumentu i zobowiązań funduszu   ← zawsze PLN; rekord oznaczany czerwono jeśli nie
  Waluta wyceny instrumentu
  Wartość instrumentu w walucie wyceny funduszu
  Procentowy udział w wartości ogółem
"""

from __future__ import annotations

import io
import os
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from typing import Any

import openpyxl
from openpyxl.styles import PatternFill

from app.services.parsers import detect_parser, parse_with_parser, is_multi_fund


# ---------------------------------------------------------------------------
# Znormalizowany wiersz
# ---------------------------------------------------------------------------

NORMALIZED_COLUMNS = [
    "Nazwa funduszu",
    "Nazwa subfunduszu",
    "Typ Funduszu",
    "Identyfikator funduszu",
    "Kod IZFiA",
    "Emitent",
    "Kraj emitenta",
    "Kod ISIN instrumentu",
    "Typ instrumentu",
    "Ilość instrumentów w portfelu",
    "Waluta wyceny instrumentu i zobowiązań funduszu",
    "Waluta wyceny instrumentu",
    "Wartość instrumentu w walucie wyceny funduszu",
    "Procentowy udział w wartości ogółem",
    "Data składu portfela",
]

RED_FILL = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")


@dataclass
class DaneRow:
    umbrella_name: str | None
    subfund_name: str | None
    fund_type: str | None
    fund_id: str | None
    izfia_id: str | None
    company_name: str
    country: str | None
    isin: str | None
    asset_type: str | None
    shares: Decimal | None
    currency_fund: str          # waluta wyceny funduszu – powinna być PLN
    currency_instrument: str    # waluta instrumentu
    value: Decimal | None
    weight_pct: Decimal | None
    snapshot_date: date | None
    currency_flag: bool = False  # True gdy currency_fund != "PLN"

    def to_dict(self) -> dict[str, Any]:
        return {
            "umbrella_name": self.umbrella_name,
            "subfund_name": self.subfund_name,
            "fund_type": self.fund_type,
            "fund_id": self.fund_id,
            "izfia_id": self.izfia_id,
            "company_name": self.company_name,
            "country": self.country,
            "isin": self.isin,
            "asset_type": self.asset_type,
            "shares": float(self.shares) if self.shares is not None else None,
            "currency_fund": self.currency_fund,
            "currency_instrument": self.currency_instrument,
            "value": float(self.value) if self.value is not None else None,
            "weight_pct": float(self.weight_pct) if self.weight_pct is not None else None,
            "snapshot_date": self.snapshot_date.isoformat() if self.snapshot_date else None,
            "currency_flag": self.currency_flag,
        }


# ---------------------------------------------------------------------------
# Parsowanie → DaneRow
# ---------------------------------------------------------------------------

def parse_file_to_dane_rows(
    file_bytes: bytes,
    filename: str,
) -> list[DaneRow]:
    """Parsuje plik portfelowy i zwraca listę znormalizowanych wierszy."""
    parser_id = detect_parser(filename, file_bytes)
    if not parser_id:
        raise ValueError(f"Nierozpoznany format pliku: {filename}")

    portfolios = parse_with_parser(parser_id, file_bytes, filename=filename, subfund_filter=None)
    if not portfolios:
        raise ValueError("Parser nie zwrócił żadnych danych.")

    rows: list[DaneRow] = []
    for portfolio in portfolios:
        # Waluta wyceny funduszu — na poziomie ParsedPortfolio
        currency_fund = (portfolio.currency or "PLN").strip().upper()

        # Typ i identyfikator funduszu — z ParsedPortfolio (ustawiony przez konwertery)
        fund_type: str | None = getattr(portfolio, "fund_type", None)
        fund_id: str | None = getattr(portfolio, "fund_id", None)
        izfia_id: str | None = getattr(portfolio, "izfia_id", None)

        for pos in portfolio.positions:
            currency_instrument = (pos.currency or "PLN").strip().upper()
            flag = currency_fund != "PLN"

            rows.append(DaneRow(
                umbrella_name=portfolio.umbrella_name,
                subfund_name=portfolio.subfund_name,
                fund_type=fund_type,
                fund_id=fund_id,
                izfia_id=izfia_id,
                company_name=pos.company_name or "",
                country=pos.country,
                isin=pos.isin,
                asset_type=pos.asset_type,
                shares=pos.shares,
                currency_fund=currency_fund,
                currency_instrument=currency_instrument,
                value=pos.value,
                weight_pct=pos.weight_pct,
                snapshot_date=portfolio.snapshot_date,
                currency_flag=flag,
            ))

    return rows


# ---------------------------------------------------------------------------
# DaneRow → xlsx
# ---------------------------------------------------------------------------

def rows_to_xlsx(rows: list[DaneRow]) -> bytes:
    """Generuje znormalizowany plik xlsx z listy DaneRow."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SkładPortfela"

    # Nagłówek
    ws.append(NORMALIZED_COLUMNS)
    header_row = ws[1]
    bold_font = openpyxl.styles.Font(bold=True)
    for cell in header_row:
        cell.font = bold_font

    # Dane
    for row in rows:
        ws.append([
            row.umbrella_name,
            row.subfund_name,
            row.fund_type,
            row.fund_id,
            row.izfia_id,
            row.company_name,
            row.country,
            row.isin,
            row.asset_type,
            float(row.shares) if row.shares is not None else None,
            row.currency_fund,
            row.currency_instrument,
            float(row.value) if row.value is not None else None,
            float(row.weight_pct) if row.weight_pct is not None else None,
            row.snapshot_date.isoformat() if row.snapshot_date else None,
        ])
        # Zaznacz czerwono wiersze z walutą funduszu != PLN
        if row.currency_flag:
            excel_row = ws[ws.max_row]
            for cell in excel_row:
                cell.fill = RED_FILL

    # Szerokości kolumn — autofit (przybliżone)
    column_widths = [20, 30, 12, 12, 10, 40, 10, 15, 25, 15, 15, 15, 20, 15, 14]
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Pomocnicze
# ---------------------------------------------------------------------------

def parsed_filename(original_filename: str) -> str:
    """Zwraca nazwę pliku z appendiksem _parsed (przed rozszerzeniem)."""
    root, ext = os.path.splitext(original_filename)
    return f"{root}_parsed.xlsx"
