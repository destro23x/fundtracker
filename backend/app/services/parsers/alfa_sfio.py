"""
Parser dla pliku ALFA SFIO PKO (format KNF regulacyjny).

Format pliku (jeden lub więcej arkuszy):
  Wiersz 1: nagłówek (18 kolumn)
  Wiersze 2+: dane — jeden wiersz = jedna pozycja

Kolumny:
  [0]  Identyfikator funduszu lub subfunduszu  — ID (np. KBC004)
  [1]  Nazwa subfunduszu                       — nazwa
  [2]  Typ funduszu                            — SFIO / FIO itd.
  [3]  Standardowy identifikator subfunduszu   — N/D lub kod
  [4]  Waluta wyceny funduszu                  — PLN
  [5]  Nazwa emitenta                          — nazwa spółki/emitenta
  [6]  Identyfikator instrumentu               — ISIN lub lokalny ID
  [7]  Alternatywny identyfikator instrumentu  — N/D lub ticker
  [8]  Typ instrumentu                         — Obligacje, Akcje…
  [9]  Kategoria instrumentu
  [10] Kraj emitenta
  [11] Kraj ryzyka
  [12] Waluta instrumentu
  [13] Ilosc instrumentow w portfelu           — ilość (sztuki / nominał)
  [14] Wartosc instrumentu w walucie wyceny funduszu — wartość w PLN
  [15] Procentowy udzial w wartosci aktywow ogolem   — ułamek 0-1
  [16] Procentowy udzial w NAV                       — ułamek 0-1
  [17] Informacje uzupelniajace

Uwaga: plik może zawierać uszkodzone referencje do obiektów rysunkowych (broken
drawing relationship). Parser naprawia to automatycznie przed wczytaniem arkusza.
"""

from __future__ import annotations

import io
import re
import zipfile
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation

import openpyxl

# ---------------------------------------------------------------------------
# Stałe kolumn
# ---------------------------------------------------------------------------

COL_FUND_ID = 0
COL_SUBFUND_NAME = 1
COL_FUND_TYPE = 2
COL_CURRENCY_FUND = 4
COL_COMPANY_NAME = 5
COL_ISIN = 6
COL_ALT_ID = 7
COL_ASSET_TYPE = 8
COL_INSTR_CURRENCY = 12
COL_SHARES = 13
COL_VALUE = 14
COL_WEIGHT_TOTAL = 15  # % aktywów ogółem
COL_WEIGHT_NAV = 16    # % NAV

EXPECTED_HEADER_COL0 = "Identyfikator funduszu lub subfunduszu"

ND_VALUES = {"n/d", "nd", "n.d.", "-", ""}

# ---------------------------------------------------------------------------
# Typy danych
# ---------------------------------------------------------------------------


@dataclass
class AlfaPosition:
    company_name: str
    isin: str | None
    currency: str
    shares: Decimal | None
    value: Decimal | None
    weight_pct: Decimal | None  # zawsze w % (0–100)
    asset_type: str | None


@dataclass
class AlfaSubfundSnapshot:
    subfund_name: str
    fund_id: str | None
    fund_type: str | None
    snapshot_date: date | None
    positions: list[AlfaPosition] = field(default_factory=list)

    @property
    def total_value(self) -> Decimal:
        return sum(
            (p.value for p in self.positions if p.value is not None),
            Decimal("0"),
        )


# ---------------------------------------------------------------------------
# Pomocnicze funkcje
# ---------------------------------------------------------------------------


def _to_decimal(val) -> Decimal | None:
    if val is None:
        return None
    try:
        return Decimal(str(val))
    except InvalidOperation:
        return None


def _nd(val) -> str | None:
    """Zwraca None jeśli wartość to 'N/D' lub pusta, inaczej strip."""
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in ND_VALUES:
        return None
    return s


def _normalize_isin(raw: str | None) -> str | None:
    if not raw:
        return None
    s = raw.strip()
    if s.lower() in ND_VALUES:
        return None
    # ISIN: 2 litery + 10 znaków alfanumerycznych
    if re.match(r"^[A-Z]{2}[A-Z0-9]{10}$", s):
        return s
    # Lokalny identyfikator (np. PLxxxx) — zachowaj
    return s


def _add_pko_prefix(name: str) -> str:
    """Dodaje prefiks 'PKO' jeśli nazwa jeszcze go nie zawiera."""
    if name.upper().startswith("PKO"):
        return name
    return f"PKO {name}"


def _extract_date_from_filename(filename: str) -> date | None:
    """Wyciąga datę w formacie YYYY-MM-DD lub YYYYMMDD z nazwy pliku."""
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", filename)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    m = re.search(r"(\d{4})(\d{2})(\d{2})", filename)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return None


def _fix_xlsx_bytes(file_bytes: bytes) -> bytes:
    """
    Naprawia uszkodzone pliki xlsx, które mają referencje do nieistniejących
    plików rysunkowych (drawing1.xml, vmlDrawing1.vml). Usuwa te relacje
    z pliku .rels, żeby openpyxl mógł wczytać plik bez błędu KeyError.
    """
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes), "r") as zin:
            names = set(zin.namelist())
            # Sprawdź czy są uszkodzone relacje
            needs_fix = False
            for n in names:
                if n.endswith(".rels"):
                    content = zin.read(n).decode("utf-8", errors="ignore")
                    if "drawing" in content.lower():
                        # Sprawdź czy referenced file istnieje
                        rel_matches = re.findall(r'Target="([^"]*drawing[^"]*)"', content, re.IGNORECASE)
                        for rel_target in rel_matches:
                            # Normalizuj ścieżkę względną
                            base = "/".join(n.split("/")[:-2]) if "/" in n else ""
                            target_path = (base + "/" + rel_target.lstrip("./")).lstrip("/")
                            if target_path not in names and rel_target.lstrip("./") not in names:
                                needs_fix = True
                                break
                if needs_fix:
                    break

            if not needs_fix:
                return file_bytes

            # Przebuduj zip bez uszkodzonych relacji
            buf = io.BytesIO()
            with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
                for name in zin.namelist():
                    data = zin.read(name)
                    if name.endswith(".rels"):
                        text = data.decode("utf-8", errors="ignore")
                        # Usuń relacje do rysunków i vml
                        text = re.sub(
                            r'<Relationship\s[^>]*(?:drawing|vmlDrawing)[^>]*/?>',
                            "",
                            text,
                            flags=re.IGNORECASE,
                        )
                        data = text.encode("utf-8")
                    zout.writestr(name, data)
            return buf.getvalue()
    except Exception:
        # Jeśli coś pójdzie nie tak — zwróć oryginał (openpyxl wyrzuci własny błąd)
        return file_bytes


# ---------------------------------------------------------------------------
# Główna funkcja parsująca
# ---------------------------------------------------------------------------


def parse_alfa_sfio(
    file_bytes: bytes,
    filename: str = "",
    subfund_filter: str | None = None,
) -> list[AlfaSubfundSnapshot]:
    """
    Parsuje plik ALFA SFIO (format KNF regulacyjny).

    Args:
        file_bytes:     Zawartość pliku .xlsx
        filename:       Nazwa pliku (do wykrycia daty)
        subfund_filter: Jeśli podany — zwraca tylko dopasowany subfundusz

    Returns:
        Lista AlfaSubfundSnapshot (po jednym na subfundusz).
    """
    fixed_bytes = _fix_xlsx_bytes(file_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(fixed_bytes), data_only=True)

    if not wb.sheetnames:
        raise ValueError("Plik jest pusty (brak arkuszy).")

    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Arkusz jest pusty.")

    # Walidacja nagłówka
    header_val = str(rows[0][COL_FUND_ID]).strip() if rows[0][COL_FUND_ID] else ""
    if header_val != EXPECTED_HEADER_COL0:
        raise ValueError(
            f"Nieoczekiwany nagłówek (kol. 0): '{header_val}'. "
            f"Oczekiwano: '{EXPECTED_HEADER_COL0}'"
        )

    snapshot_date = _extract_date_from_filename(filename)

    subfunds: dict[str, AlfaSubfundSnapshot] = {}

    for row in rows[1:]:
        if not row or row[COL_SUBFUND_NAME] is None:
            continue

        subfund_name = _add_pko_prefix(str(row[COL_SUBFUND_NAME]).strip())
        if not subfund_name:
            continue

        # Filtr
        if subfund_filter and subfund_filter.lower() not in subfund_name.lower():
            continue

        # Utwórz snapshot jeśli nieznany
        if subfund_name not in subfunds:
            fund_id = _nd(row[COL_FUND_ID])
            fund_type = _nd(row[COL_FUND_TYPE])
            subfunds[subfund_name] = AlfaSubfundSnapshot(
                subfund_name=subfund_name,
                fund_id=fund_id,
                fund_type=fund_type,
                snapshot_date=snapshot_date,
            )

        # --- Parsuj pozycję ---
        company_name = _nd(row[COL_COMPANY_NAME])
        isin_raw = _nd(row[COL_ISIN])
        asset_type = _nd(row[COL_ASSET_TYPE])
        currency = _nd(row[COL_INSTR_CURRENCY]) or _nd(row[COL_CURRENCY_FUND]) or "PLN"

        shares = _to_decimal(row[COL_SHARES])
        value = _to_decimal(row[COL_VALUE])

        # Udział w NAV — ułamek 0-1 → konwertujemy na %
        weight_raw = _to_decimal(row[COL_WEIGHT_NAV])
        if weight_raw is None:
            weight_raw = _to_decimal(row[COL_WEIGHT_TOTAL])
        weight_pct = weight_raw * Decimal("100") if weight_raw is not None else None

        # Nazwa wyświetlana: emitent > ISIN > typ instrumentu
        display_name = (
            company_name
            or _normalize_isin(isin_raw)
            or asset_type
            or "Nieznany instrument"
        )

        pos = AlfaPosition(
            company_name=display_name[:500],
            isin=_normalize_isin(isin_raw),
            currency=currency,
            shares=shares,
            value=value,
            weight_pct=weight_pct,
            asset_type=asset_type[:50] if asset_type else None,
        )
        subfunds[subfund_name].positions.append(pos)

    return list(subfunds.values())


def list_subfunds(file_bytes: bytes) -> list[str]:
    """Zwraca listę nazw subfunduszy bez pełnego parsowania."""
    fixed_bytes = _fix_xlsx_bytes(file_bytes)
    wb = openpyxl.load_workbook(io.BytesIO(fixed_bytes), data_only=True)
    if not wb.sheetnames:
        return []
    ws = wb.worksheets[0]
    names: list[str] = []
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[COL_SUBFUND_NAME] is None:
            continue
        name = _add_pko_prefix(str(row[COL_SUBFUND_NAME]).strip())
        if name and name not in seen:
            names.append(name)
            seen.add(name)
    return names
