"""
Parser dla pliku Alior TFI SFIO (format KNF regulacyjny).

Format pliku (arkusz 'DANE'):
  Wiersz 1: tytuł (pomijany)
  Wiersz 2: nagłówki (17 kolumn)
  Wiersze 3+: dane — jeden wiersz = jedna pozycja

Kolumny:
  [0]  Identyfikator IZFIA funduszu lub subfunduszu  — np. ALR010
  [1]  Nazwa funduszu                               — ALIOR SFIO
  [2]  Nazwa subfunduszu                            — ALIOR Akcji Polskich
  [3]  Inny Identyfikator funduszu                  — ISIN parasola
  [4]  Typ funduszu                                 — SFIO
  [5]  Emitent                                      — nazwa spółki
  [6]  Kod ISIN instrumentu                         — ISIN akcji/obligacji
  [7]  Typ instrumentu                              — Akcje, Obligacje…
  [8]  Kraj emitenta                                — PL, US…
  [9]  Waluta wyceny instrumentu                    — PLN, USD, EUR…
  [10] Ilość instrumentów w portfelu                — ilość (może być ułamkowa)
  [11] Wartość instrumentu w walucie wyceny funduszu— wartość w PLN
  [12] Udział procentowy w aktywach                 — ułamek dziesiętny (0.038…) → ×100 = %
  [13] Waluta wyceny aktywów i zobowiązań funduszu  — PLN
  [14] Inny standardowy identyfikator instrumentu  — alt ID lub N/D
  [15] Kategoria instrumentu
  [16] Informacje uzupełniające
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation

import openpyxl

# ---------------------------------------------------------------------------
# Stałe kolumn
# ---------------------------------------------------------------------------

COL_IZFIA_ID      = 0
COL_FUND_NAME     = 1
COL_SUBFUND_NAME  = 2
COL_FUND_ISIN     = 3
COL_FUND_TYPE     = 4
COL_COMPANY_NAME  = 5
COL_ISIN          = 6
COL_ASSET_TYPE    = 7
COL_COUNTRY       = 8
COL_CURRENCY_INSTR= 9
COL_SHARES        = 10
COL_VALUE         = 11
COL_WEIGHT_PCT    = 12
COL_CURRENCY_FUND = 13
COL_ALT_ID        = 14

EXPECTED_HEADER_COL0 = "Identyfikator IZFIA funduszu lub subfunduszu"
SHEET_NAME = "DANE"

ND_VALUES = {"n/d", "nd", "n.d.", "-", "", "none", "null"}


# ---------------------------------------------------------------------------
# Typy danych
# ---------------------------------------------------------------------------

@dataclass
class AliorPosition:
    company_name: str
    isin: str | None
    currency: str
    shares: Decimal | None
    value: Decimal | None
    weight_pct: Decimal | None
    asset_type: str | None
    country: str | None = None


@dataclass
class AliorSubfundSnapshot:
    subfund_name: str
    fund_name: str | None
    izfia_id: str | None
    izfia_code: str | None          # Kod IZFiA (kol. 0), np. ALR010
    fund_type: str | None
    snapshot_date: date | None
    positions: list[AliorPosition] = field(default_factory=list)

    @property
    def total_value(self) -> Decimal:
        return sum(
            (p.value for p in self.positions if p.value is not None),
            Decimal("0"),
        )


# ---------------------------------------------------------------------------
# Pomocnicze
# ---------------------------------------------------------------------------

def _to_decimal(val) -> Decimal | None:
    if val is None:
        return None
    try:
        return Decimal(str(val))
    except InvalidOperation:
        return None


def _nd(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in ND_VALUES:
        return None
    return s


def _normalize_isin(raw: str | None) -> str | None:
    if not raw:
        return None
    s = raw.strip().upper()
    if s.lower() in ND_VALUES:
        return None
    if re.match(r"^[A-Z]{2}[A-Z0-9]{10}$", s):
        return s
    if len(s) >= 6:
        return s
    return None


def _extract_date_from_filename(filename: str) -> date | None:
    # YYYY-MM-DD / YYYYMMDD
    m = re.search(r"(\d{4})[._-](\d{2})[._-](\d{2})", filename)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    # DDMMYYYY (bez separatorów, np. 30042026)
    m = re.search(r"(\d{2})(\d{2})(\d{4})(?!\d)", filename)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    # DD-MM-YYYY
    m = re.search(r"(\d{2})[._-](\d{2})[._-](\d{4})", filename)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    return None


# ---------------------------------------------------------------------------
# Główna funkcja parsująca
# ---------------------------------------------------------------------------

def parse_alior_sfio(
    file_bytes: bytes,
    filename: str = "",
    subfund_filter: str | None = None,
) -> list[AliorSubfundSnapshot]:
    """
    Parsuje plik Alior TFI SFIO (format KNF).

    Returns:
        Lista AliorSubfundSnapshot (po jednym na subfundusz).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    if not wb.sheetnames:
        raise ValueError("Plik jest pusty (brak arkuszy).")

    sheet_name = SHEET_NAME if SHEET_NAME in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        raise ValueError("Arkusz jest pusty lub ma za mało wierszy.")

    # Wiersz 0 = tytuł, wiersz 1 = nagłówki — sprawdź nagłówek
    header_row = rows[1]
    header_val = str(header_row[COL_IZFIA_ID]).strip() if header_row[COL_IZFIA_ID] else ""
    if header_val != EXPECTED_HEADER_COL0:
        raise ValueError(
            f"Nieoczekiwany nagłówek (kol. 0): '{header_val}'. "
            f"Oczekiwano: '{EXPECTED_HEADER_COL0}'"
        )

    snapshot_date = _extract_date_from_filename(filename)

    # Fallback: wyciągnij datę z tytułu w wierszu 0 (np. "na dzień 30/04/2026")
    if snapshot_date is None and rows:
        title_text = str(rows[0][1] or rows[0][0] or "").strip()
        m = re.search(r"(\d{1,2})/(\d{2})/(\d{4})", title_text)
        if m:
            try:
                snapshot_date = date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            except ValueError:
                pass

    subfunds: dict[str, AliorSubfundSnapshot] = {}

    for row in rows[2:]:          # dane od wiersza 3 (index 2)
        if not row or row[COL_SUBFUND_NAME] is None:
            continue

        subfund_name = str(row[COL_SUBFUND_NAME]).strip()
        if not subfund_name:
            continue

        if subfund_filter and subfund_filter.lower() not in subfund_name.lower():
            continue

        if subfund_name not in subfunds:
            subfunds[subfund_name] = AliorSubfundSnapshot(
                subfund_name=subfund_name,
                fund_name=_nd(row[COL_FUND_NAME]),
                izfia_id=_nd(row[COL_FUND_ISIN]) if len(row) > COL_FUND_ISIN else _nd(row[COL_IZFIA_ID]),
                izfia_code=_nd(row[COL_IZFIA_ID]),
                fund_type=_nd(row[COL_FUND_TYPE]),
                snapshot_date=snapshot_date,
            )

        company_name = _nd(row[COL_COMPANY_NAME])
        isin_raw = _nd(row[COL_ISIN])
        if not isin_raw:
            isin_raw = _nd(row[COL_ALT_ID])

        asset_type = _nd(row[COL_ASSET_TYPE])
        currency = _nd(row[COL_CURRENCY_INSTR]) or _nd(row[COL_CURRENCY_FUND]) or "PLN"

        shares = _to_decimal(row[COL_SHARES])
        value = _to_decimal(row[COL_VALUE])

        # Udział: ułamek dziesiętny (0.038…) → procenty
        weight_raw = _to_decimal(row[COL_WEIGHT_PCT])
        weight_pct = (weight_raw * 100).quantize(Decimal("0.0001")) if weight_raw is not None else None

        display_name = (
            company_name
            or _normalize_isin(isin_raw)
            or asset_type
            or "Nieznany instrument"
        )

        pos = AliorPosition(
            company_name=display_name[:500],
            isin=_normalize_isin(isin_raw),
            currency=currency,
            shares=shares,
            value=value,
            weight_pct=weight_pct,
            asset_type=asset_type[:50] if asset_type else None,
            country=_nd(row[COL_COUNTRY]) if len(row) > COL_COUNTRY else None,
        )
        subfunds[subfund_name].positions.append(pos)

    return list(subfunds.values())


def list_subfunds(file_bytes: bytes) -> list[str]:
    """Zwraca listę nazw subfunduszy bez pełnego parsowania."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if not wb.sheetnames:
        return []
    sheet_name = SHEET_NAME if SHEET_NAME in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    names: list[str] = []
    seen: set[str] = set()
    # Dane od wiersza 3 (min_row=3)
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or row[COL_SUBFUND_NAME] is None:
            continue
        name = str(row[COL_SUBFUND_NAME]).strip()
        if name and name not in seen:
            names.append(name)
            seen.add(name)
    return names
