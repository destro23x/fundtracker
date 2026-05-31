"""
Parser dedykowany dla plików Goldman Sachs TFI.

Format pliku (jeden arkusz Sheet1):
  Wiersz 0: nagłówek
  Wiersze 1+: dane — jeden wiersz = jedna pozycja

Kolumny (indeks 0-15):
  [0]  Nazwa Funduszu / Nazwa Subfunduszu  — nazwa subfunduszu
  [1]  Nazwa Parasola                      — nazwa parasola (umbrella)
  [2]  Typ                                 — typ funduszu (SFIO, FIO…)
  [3]  KNF_ID
  [4]  IZFIA_ID
  [5]  Kategoria / Typ Instrumentu         — typ aktywa (Akcje, Obligacje…)
  [6]  ISIN                                — kod ISIN lub lokalny identyfikator
  [7]  NAZWA SKRÓCONA INSTRUMENTU          — ticker / skrót
  [8]  NAZWA PEŁNA INSTRUMENTU             — pełna nazwa spółki
  [9]  WALUTA                              — waluta instrumentu
  [10] Nazwa Emitenta                      — pełna nazwa emitenta
  [11] Kod Kraju Emitenta
  [12] ILOŚĆ                               — liczba jednostek
  [13] WARTOŚĆ CAŁKOWITA                   — wartość w walucie wyceny (PLN)
  [14] WALUTA WYCENY                       — zawsze PLN
  [15] UDZIAŁ W AKTYWACH                   — udział 0-1 (może być ujemny dla pochodnych)

Plik zawiera pozycje wielu subfunduszy jednocześnie — można sparsować wszystkie
lub wybrać konkretny subfundusz po nazwie.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl

# ---------------------------------------------------------------------------
# Stałe kolumn
# ---------------------------------------------------------------------------

COL_SUBFUND_NAME = 0
COL_UMBRELLA = 1
COL_FUND_TYPE = 2
COL_KNF_ID = 3
COL_ASSET_TYPE = 5
COL_ISIN = 6
COL_SHORT_NAME = 7
COL_FULL_NAME = 8
COL_CURRENCY = 9
COL_COUNTRY = 11
COL_SHARES = 12
COL_VALUE = 13
COL_VALUATION_CURRENCY = 14
COL_WEIGHT = 15

EXPECTED_HEADER = "Nazwa Funduszu / Nazwa Subfunduszu"

# Mapowanie typów aktywów na ujednolicone wartości
ASSET_TYPE_MAP = {
    "akcje": "stock",
    "obligacje skarbowe": "bond_government",
    "obligacje korporacyjne": "bond_corporate",
    "hipoteczne listy zastawne": "covered_bond",
    "obligacje emitowane przez jednostki samorządu terytorialnego": "bond_municipal",
    "tytuły uczestnictwa zagranicznego": "etf_foreign",
    "fundusze inwestycyjne": "fund",
    "waluta": "cash",
    "trankacja na obligacji typu buy-sell-back": "repo",
    "instrumenty pochodne - fx forward": "derivative_fx",
    "instrumenty pochodne - swap": "derivative_swap",
    "instrumenty pochodne - kontrakty terminowe indeksy giełdowe akcji": "derivative_futures_index",
    "instrumenty pochodne - kontrakty terminowe obligacje": "derivative_futures_bond",
    "instrumenty pochodne - kontrakty terminowe akcje": "derivative_futures_equity",
}


# ---------------------------------------------------------------------------
# Typy danych
# ---------------------------------------------------------------------------

@dataclass
class GoldmanPosition:
    company_name: str
    short_name: str | None
    isin: str | None
    currency: str
    country_code: str | None
    shares: Decimal | None
    value: Decimal | None        # wartość w PLN
    weight_pct: Decimal | None   # udział % (może być ujemny dla pochodnych)
    asset_type_raw: str | None   # oryginalna kategoria z pliku
    asset_type: str | None       # ujednolicona kategoria


@dataclass
class GoldmanSubfundSnapshot:
    subfund_name: str
    umbrella_name: str | None
    fund_type: str | None
    knf_id: str | None
    snapshot_date: date | None
    positions: list[GoldmanPosition] = field(default_factory=list)

    @property
    def total_value(self) -> Decimal:
        """Suma wartości pozycji w PLN."""
        return sum(
            (p.value for p in self.positions if p.value is not None),
            Decimal("0"),
        )


# ---------------------------------------------------------------------------
# Pomocnicze funkcje
# ---------------------------------------------------------------------------

def _to_decimal(val) -> Decimal | None:
    if val is None:
        return None
    try:
        return Decimal(str(val))
    except InvalidOperation:
        return None


def _normalize_isin(raw: str | None) -> str | None:
    """Zwraca wartość jeżeli wygląda jak ISIN (12 znaków alfanumerycznych, zaczyna się od 2 liter)."""
    if not raw:
        return None
    s = str(raw).strip()
    if re.match(r"^[A-Z]{2}[A-Z0-9]{10}$", s):
        return s
    return s  # zostawiamy nawet jeśli nie jest ISIN — to lokalny identyfikator


def _map_asset_type(raw: str | None) -> str | None:
    if not raw:
        return None
    return ASSET_TYPE_MAP.get(raw.strip().lower(), "other")


def _extract_date_from_filename(filename: str) -> date | None:
    """
    Wyciąga datę z nazwy pliku.
    Przykład: pl_sklady_portfeli_funduszy_goldman_sachs_tfi_-_2026-03-31.xlsx
    """
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", filename)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    # Fallback: YYYYMMDD
    m = re.search(r"(\d{4})(\d{2})(\d{2})", filename)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return None


def _validate_header(row: tuple) -> bool:
    return bool(row) and str(row[COL_SUBFUND_NAME]).strip() == EXPECTED_HEADER


# ---------------------------------------------------------------------------
# Główna funkcja parsująca
# ---------------------------------------------------------------------------

def parse_goldman_sachs(
    file_bytes: bytes,
    filename: str = "",
    subfund_filter: str | None = None,
) -> list[GoldmanSubfundSnapshot]:
    """
    Parsuje plik Goldman Sachs TFI i zwraca listę snapshotów subfunduszy.

    Args:
        file_bytes:      Zawartość pliku .xlsx
        filename:        Nazwa pliku (używana do wykrycia daty)
        subfund_filter:  Jeśli podany — zwraca tylko ten subfundusz (dopasowanie
                         częściowe, case-insensitive). Np. "Akcji Polskich Plus".

    Returns:
        Lista GoldmanSubfundSnapshot — po jednym obiekcie na subfundusz.
        Jeśli subfund_filter jest podany i nie znajdzie subfunduszu, lista jest pusta.
    """
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    if "Sheet1" not in wb.sheetnames:
        raise ValueError(
            f"Oczekiwano arkusza 'Sheet1'. Dostępne: {wb.sheetnames}"
        )

    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Plik jest pusty.")

    if not _validate_header(rows[0]):
        raise ValueError(
            f"Nieoczekiwany nagłówek. Pierwsza kolumna: '{rows[0][0]}'. "
            f"Oczekiwano: '{EXPECTED_HEADER}'"
        )

    snapshot_date = _extract_date_from_filename(filename)

    # Grupuj wiersze po nazwie subfunduszu
    subfunds: dict[str, GoldmanSubfundSnapshot] = {}

    for row_idx, row in enumerate(rows[1:], start=2):
        # Pomiń puste wiersze
        if not row or row[COL_SUBFUND_NAME] is None:
            continue

        subfund_name = str(row[COL_SUBFUND_NAME]).strip()
        if not subfund_name:
            continue

        # Filtr po nazwie subfunduszu
        if subfund_filter and subfund_filter.lower() not in subfund_name.lower():
            continue

        # Utwórz lub pobierz obiekt subfunduszu
        if subfund_name not in subfunds:
            umbrella = str(row[COL_UMBRELLA]).strip() if row[COL_UMBRELLA] else None
            fund_type = str(row[COL_FUND_TYPE]).strip() if row[COL_FUND_TYPE] else None
            knf_id = str(row[COL_KNF_ID]).strip() if row[COL_KNF_ID] else None
            subfunds[subfund_name] = GoldmanSubfundSnapshot(
                subfund_name=subfund_name,
                umbrella_name=umbrella,
                fund_type=fund_type,
                knf_id=knf_id,
                snapshot_date=snapshot_date,
            )

        # Parsuj pozycję
        asset_type_raw = str(row[COL_ASSET_TYPE]).strip() if row[COL_ASSET_TYPE] else None
        isin_raw = str(row[COL_ISIN]).strip() if row[COL_ISIN] else None
        short_name = str(row[COL_SHORT_NAME]).strip() if row[COL_SHORT_NAME] else None
        full_name = str(row[COL_FULL_NAME]).strip() if row[COL_FULL_NAME] else None
        currency = str(row[COL_CURRENCY]).strip() if row[COL_CURRENCY] else "PLN"
        country = str(row[COL_COUNTRY]).strip() if row[COL_COUNTRY] else None

        shares = _to_decimal(row[COL_SHARES])
        value = _to_decimal(row[COL_VALUE])

        # UDZIAŁ W AKTYWACH jest na skali 0-1, konwertujemy na 0-100%
        weight_raw = _to_decimal(row[COL_WEIGHT])
        weight_pct = weight_raw * Decimal("100") if weight_raw is not None else None

        # Nazwa wyświetlana: pełna > skrócona
        company_name = full_name or short_name or isin_raw or "Nieznany instrument"

        pos = GoldmanPosition(
            company_name=company_name,
            short_name=short_name,
            isin=_normalize_isin(isin_raw),
            currency=currency,
            country_code=country,
            shares=shares,
            value=value,
            weight_pct=weight_pct,
            asset_type_raw=asset_type_raw,
            asset_type=_map_asset_type(asset_type_raw),
        )
        subfunds[subfund_name].positions.append(pos)

    return list(subfunds.values())


def list_subfunds(file_bytes: bytes) -> list[str]:
    """
    Zwraca listę nazw wszystkich subfunduszy w pliku (bez pełnego parsowania).
    Przydatne do pokazania użytkownikowi wyboru przed uploadem.
    """
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if "Sheet1" not in wb.sheetnames:
        return []
    ws = wb["Sheet1"]
    seen: list[str] = []
    seen_set: set[str] = set()
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row and row[COL_SUBFUND_NAME]:
            name = str(row[COL_SUBFUND_NAME]).strip()
            if name and name not in seen_set:
                seen.append(name)
                seen_set.add(name)
    return seen
