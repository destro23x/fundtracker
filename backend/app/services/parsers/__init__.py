"""
Rejestr parserów — mapuje rozpoznany format pliku na konkretny parser.

Każdy parser musi zwracać obiekt kompatybilny z ParsedPortfolio z excel_parser.py
(lub listę takich obiektów dla plików multi-funduszowych).

Dodawanie nowego parsera:
  1. Stwórz plik w tym katalogu, np. pekao_tfi.py
  2. Zaimplementuj funkcję detect(filename, first_row) -> bool
  3. Dodaj go do PARSERS poniżej
"""

from __future__ import annotations

from dataclasses import dataclass, field
from decimal import Decimal
from datetime import date
import re
import io

from app.services.excel_parser import ParsedPortfolio, ParsedPosition


# ---------------------------------------------------------------------------
# Konwersja z GoldmanPosition / GoldmanSubfundSnapshot → ParsedPortfolio
# ---------------------------------------------------------------------------

def _alfa_to_parsed(snapshot) -> ParsedPortfolio:
    """Konwertuje AlfaSubfundSnapshot na uniwersalny ParsedPortfolio."""
    result = ParsedPortfolio()
    result.snapshot_date = snapshot.snapshot_date
    result.total_value = snapshot.total_value if snapshot.total_value else None
    result.currency = "PLN"
    result.subfund_name = snapshot.subfund_name
    result.fund_type = getattr(snapshot, "fund_type", None)
    result.fund_id = getattr(snapshot, "fund_id", None)  # Alfa używa fund_id

    for p in snapshot.positions:
        result.positions.append(
            ParsedPosition(
                company_name=p.company_name,
                ticker=None,
                isin=p.isin,
                shares=p.shares,
                value=p.value,
                weight_pct=p.weight_pct,
                currency=p.currency or "PLN",
                asset_type=p.asset_type,
            )
        )
    return result


def _goldman_to_parsed(snapshot) -> ParsedPortfolio:
    """Konwertuje GoldmanSubfundSnapshot na uniwersalny ParsedPortfolio."""

    result = ParsedPortfolio()
    result.snapshot_date = snapshot.snapshot_date
    result.total_value = snapshot.total_value if snapshot.total_value else None
    result.currency = "PLN"
    result.subfund_name = snapshot.subfund_name
    result.umbrella_name = snapshot.umbrella_name  # Nazwa Parasola → Fundusz
    result.fund_type = getattr(snapshot, "fund_type", None)
    result.fund_id = getattr(snapshot, "knf_id", None)
    result.raw_headers = [
        "Nazwa Funduszu", "Nazwa Parasola", "Typ", "KNF_ID", "IZFIA_ID",
        "Kategoria", "ISIN", "Nazwa skrócona", "Nazwa pełna", "Waluta",
        "Emitent", "Kraj", "Ilość", "Wartość całkowita", "Waluta wyceny",
        "Udział w aktywach",
    ]

    for p in snapshot.positions:
        result.positions.append(
            ParsedPosition(
                company_name=(p.company_name or "")[:500],
                ticker=(p.short_name or "")[:255] or None,
                isin=p.isin,
                shares=p.shares,
                value=p.value,
                weight_pct=p.weight_pct,
                currency=p.currency or "PLN",
                asset_type=(p.asset_type or "")[:50] or None,
                country=p.country_code,
            )
        )
    return result


def _noble_to_parsed(snapshot) -> ParsedPortfolio:
    """Konwertuje NobleSubfundSnapshot na uniwersalny ParsedPortfolio."""
    result = ParsedPortfolio()
    result.snapshot_date = snapshot.snapshot_date
    result.total_value = snapshot.total_value if snapshot.total_value else None
    result.currency = "PLN"
    result.subfund_name = snapshot.subfund_name
    result.umbrella_name = snapshot.fund_name or None
    result.fund_type = getattr(snapshot, "fund_type", None)
    result.fund_id = getattr(snapshot, "izfia_id", None)  # Noble używa izfia_id
    result.izfia_id = getattr(snapshot, "izfia_code", None)  # Kod IZFiA (kol. 0)

    for p in snapshot.positions:
        result.positions.append(
            ParsedPosition(
                company_name=p.company_name,
                ticker=None,
                isin=p.isin,
                shares=p.shares,
                value=p.value,
                weight_pct=None,
                currency=p.currency or "PLN",
                asset_type=p.asset_type,
                country=getattr(p, "country", None),
            )
        )
    return result


def _alior_to_parsed(snapshot) -> ParsedPortfolio:
    """Konwertuje AliorSubfundSnapshot na uniwersalny ParsedPortfolio."""
    result = ParsedPortfolio()
    result.snapshot_date = snapshot.snapshot_date
    result.total_value = snapshot.total_value if snapshot.total_value else None
    result.currency = "PLN"
    result.subfund_name = snapshot.subfund_name
    result.umbrella_name = snapshot.fund_name  # np. "ALIOR SFIO"
    result.fund_type = getattr(snapshot, "fund_type", None)
    result.fund_id = getattr(snapshot, "izfia_id", None)  # Alior używa izfia_id
    result.izfia_id = getattr(snapshot, "izfia_code", None)  # Kod IZFiA (kol. 0)

    for p in snapshot.positions:
        result.positions.append(
            ParsedPosition(
                company_name=p.company_name,
                ticker=None,
                isin=p.isin,
                shares=p.shares,
                value=p.value,
                weight_pct=p.weight_pct,
                currency=p.currency or "PLN",
                asset_type=p.asset_type,
                country=getattr(p, "country", None),
            )
        )
    return result


def _pzu_to_parsed(snapshot) -> ParsedPortfolio:
    """Konwertuje PzuSubfundSnapshot na uniwersalny ParsedPortfolio."""
    result = ParsedPortfolio()
    result.snapshot_date = snapshot.snapshot_date
    result.total_value = snapshot.total_value if snapshot.total_value else None
    result.currency = getattr(snapshot, "currency_fund", None) or "PLN"
    result.subfund_name = snapshot.subfund_name
    result.umbrella_name = snapshot.fund_name or None
    result.fund_type = getattr(snapshot, "fund_type", None)
    result.fund_id = getattr(snapshot, "fund_id", None)
    result.izfia_id = getattr(snapshot, "izfia_code", None)  # Kod IZFiA (kol. 0)

    for p in snapshot.positions:
        result.positions.append(
            ParsedPosition(
                company_name=p.company_name,
                ticker=None,
                isin=p.isin,
                shares=p.shares,
                value=p.value,
                weight_pct=p.weight_pct,
                currency=p.currency or "PLN",
                asset_type=p.asset_type,
                country=getattr(p, "country", None),
            )
        )
    return result


def _uniqa_to_parsed(snapshot) -> ParsedPortfolio:
    """Konwertuje UniqaSubfundSnapshot na uniwersalny ParsedPortfolio."""
    from app.services.parsers.uniqa_fio import FUND_FULL_NAME as UNIQA_FUND_NAME
    result = ParsedPortfolio()
    result.snapshot_date = snapshot.snapshot_date
    result.total_value = snapshot.total_value if snapshot.total_value else None
    result.currency = "PLN"
    result.umbrella_name = UNIQA_FUND_NAME
    result.subfund_name = snapshot.subfund_name
    result.fund_type = "FIO"
    result.fund_id = None  # UNIQA FIO PDF nie zawiera ISIN-type identyfikatora funduszu
    result.izfia_id = getattr(snapshot, "izfia_code", None)  # Kod IZFiA (np. AXA002)

    for p in snapshot.positions:
        result.positions.append(
            ParsedPosition(
                company_name=p.company_name,
                ticker=None,
                isin=p.isin,
                shares=p.shares,
                value=p.value,
                weight_pct=p.weight_pct,
                currency=p.currency or "PLN",
                asset_type=p.asset_type,
                country=getattr(p, "country", None),
            )
        )
    return result


def _superfund_to_parsed(snapshot) -> ParsedPortfolio:
    """Konwertuje SuperfundSubfundSnapshot na uniwersalny ParsedPortfolio."""
    result = ParsedPortfolio()
    result.snapshot_date = snapshot.snapshot_date
    result.total_value = snapshot.total_value if snapshot.total_value else None
    result.currency = "PLN"
    result.umbrella_name = "SUPERFUND"
    result.subfund_name = snapshot.subfund_name

    for p in snapshot.positions:
        result.positions.append(
            ParsedPosition(
                company_name=p.company_name,
                ticker=None,
                isin=p.isin,
                shares=p.shares,
                value=p.value,
                weight_pct=p.weight_pct,
                currency=p.currency or "PLN",
                asset_type=p.asset_type,
            )
        )
    return result


def _superfund_xlsx_to_parsed(snapshot) -> ParsedPortfolio:
    """Konwertuje SuperfundXlsxSubfundSnapshot na uniwersalny ParsedPortfolio."""
    from app.services.parsers.superfund_xlsx import UMBRELLA_NAME as SUPERFUND_NAME
    result = ParsedPortfolio()
    result.snapshot_date = snapshot.snapshot_date
    result.total_value = snapshot.total_value if snapshot.total_value else None
    result.currency = "PLN"
    result.umbrella_name = SUPERFUND_NAME  # zawsze "SUPERFUND"
    result.subfund_name = snapshot.subfund_name
    result.fund_type = None
    result.fund_id = None

    for p in snapshot.positions:
        result.positions.append(
            ParsedPosition(
                company_name=p.company_name,
                ticker=None,
                isin=p.isin,
                shares=p.shares,
                value=p.value,
                weight_pct=p.weight_pct,
                currency=p.currency or "PLN",
                asset_type=p.asset_type,
            )
        )
    return result


def _pekao_to_parsed(snapshot) -> ParsedPortfolio:
    """Konwertuje PekaoSubfundSnapshot na uniwersalny ParsedPortfolio."""
    from app.services.parsers.pekao_tfi import UMBRELLA_NAME as PEKAO_UMBRELLA
    result = ParsedPortfolio()
    result.snapshot_date = snapshot.snapshot_date
    result.total_value = snapshot.total_value if snapshot.total_value else None
    result.currency = snapshot.currency_fund or "PLN"
    result.umbrella_name = PEKAO_UMBRELLA
    result.subfund_name = snapshot.subfund_name
    result.fund_type = getattr(snapshot, "fund_type", None)
    result.fund_id = getattr(snapshot, "fund_id", None)      # ISIN subfunduszu
    result.izfia_id = getattr(snapshot, "izfia_code", None)  # identyfikator IZFiA

    for p in snapshot.positions:
        result.positions.append(
            ParsedPosition(
                company_name=p.company_name,
                ticker=None,
                isin=p.isin,
                shares=p.shares,
                value=p.value,
                weight_pct=p.weight_pct,
                currency=p.currency or result.currency,
                asset_type=p.asset_type,
                country=getattr(p, "country", None),
            )
        )
    return result


def _uniqa_xlsx_to_parsed(snapshot) -> ParsedPortfolio:
    """Konwertuje UniqaXlsxSubfundSnapshot na uniwersalny ParsedPortfolio."""
    result = ParsedPortfolio()
    result.snapshot_date = snapshot.snapshot_date
    result.total_value = snapshot.total_value if snapshot.total_value else None
    result.currency = "PLN"
    result.subfund_name = snapshot.subfund_name
    result.umbrella_name = snapshot.umbrella_name

    for p in snapshot.positions:
        result.positions.append(
            ParsedPosition(
                company_name=p.company_name,
                ticker=None,
                isin=p.isin,
                shares=p.shares,
                value=p.value,
                weight_pct=p.weight_pct,
                currency=p.currency or "PLN",
                asset_type=p.asset_type,
            )
        )
    return result


def _pko_to_parsed(snapshot) -> ParsedPortfolio:
    """Konwertuje PkoSubfundSnapshot na uniwersalny ParsedPortfolio."""
    result = ParsedPortfolio()
    result.snapshot_date = snapshot.snapshot_date
    result.total_value = None
    result.currency = snapshot.currency_fund or "PLN"
    result.subfund_name = snapshot.subfund_name
    result.umbrella_name = snapshot.umbrella_name
    result.fund_type = snapshot.fund_type
    result.fund_id = None  # PKO col 0 to kod IZFiA, nie fund_id
    result.izfia_id = getattr(snapshot, "izfia_code", None)  # np. "PKO064"

    for p in snapshot.positions:
        result.positions.append(
            ParsedPosition(
                company_name=p.company_name,
                ticker=None,
                isin=p.isin,
                shares=p.shares,
                value=p.value,
                weight_pct=p.weight_pct,
                currency=p.currency or "PLN",
                asset_type=p.asset_type,
                country=p.country,
            )
        )
    return result


def _generali_to_parsed(snapshot) -> ParsedPortfolio:
    """Konwertuje GeneraliSubfundSnapshot na uniwersalny ParsedPortfolio."""
    result = ParsedPortfolio()
    result.snapshot_date = snapshot.snapshot_date
    result.total_value = snapshot.total_value
    result.currency = snapshot.currency_fund or "PLN"
    result.subfund_name = snapshot.subfund_name
    result.umbrella_name = snapshot.umbrella_name
    result.fund_type = snapshot.fund_type
    result.fund_id = snapshot.fund_id

    for p in snapshot.positions:
        result.positions.append(
            ParsedPosition(
                company_name=p.company_name,
                ticker=None,
                isin=p.isin,
                shares=p.shares,
                value=p.value,
                weight_pct=p.weight_pct,
                currency=p.currency or "PLN",
                asset_type=p.asset_type,
                country=p.country,
            )
        )
    return result


def _bnp_to_parsed(snapshot) -> ParsedPortfolio:
    """Konwertuje BnpSubfundSnapshot na uniwersalny ParsedPortfolio."""
    result = ParsedPortfolio()
    result.snapshot_date = snapshot.snapshot_date
    result.total_value = snapshot.total_value
    result.currency = snapshot.currency_fund or "PLN"
    result.subfund_name = snapshot.subfund_name
    result.umbrella_name = snapshot.umbrella_name
    result.fund_type = snapshot.fund_type
    result.fund_id = None  # BNP col 0 to kod IZFiA, nie fund_id
    result.izfia_id = getattr(snapshot, "izfia_code", None)  # np. "PLFIO000001"

    for p in snapshot.positions:
        result.positions.append(
            ParsedPosition(
                company_name=p.company_name,
                ticker=None,
                isin=p.isin,
                shares=p.shares,
                value=p.value,
                weight_pct=p.weight_pct,
                currency=p.currency or "PLN",
                asset_type=p.asset_type,
                country=p.country,
            )
        )
    return result


def _erste_to_parsed(snapshot) -> ParsedPortfolio:
    """Konwertuje ErsteSubfundSnapshot na uniwersalny ParsedPortfolio."""
    result = ParsedPortfolio()
    result.snapshot_date = snapshot.snapshot_date
    result.total_value = snapshot.total_value if snapshot.total_value else None
    result.currency = "PLN"
    result.subfund_name = snapshot.subfund_name
    result.umbrella_name = snapshot.umbrella_name
    result.fund_type = getattr(snapshot, "fund_type", None)
    result.fund_id = getattr(snapshot, "fund_id", None)

    for p in snapshot.positions:
        result.positions.append(
            ParsedPosition(
                company_name=p.company_name,
                ticker=None,
                isin=p.isin,
                shares=p.shares,
                value=p.value,
                weight_pct=p.weight_pct,
                currency=p.currency or "PLN",
                asset_type=p.asset_type,
                country=getattr(p, "country", None),
            )
        )
    return result


# ---------------------------------------------------------------------------
# Wykrywanie formatu PDF
# ---------------------------------------------------------------------------

def _detect_pdf_parser(file_bytes: bytes) -> str | None:
    """Wykrywa parser dla pliku PDF na podstawie tekstu pierwszej strony."""
    try:
        import pdfplumber
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            if not pdf.pages:
                return None
            text = pdf.pages[0].extract_text() or ""
            first_line = text.split("\n")[0].strip() if text else ""

            from app.services.parsers.uniqa_fio import DETECTION_TEXT as UNIQA_TEXT
            from app.services.parsers.superfund_tfi import DETECTION_TEXT as SUPERFUND_TEXT

            if UNIQA_TEXT in text:
                return "uniqa_fio"
            if SUPERFUND_TEXT in first_line or SUPERFUND_TEXT in text:
                return "superfund_tfi"
    except Exception:
        pass
    return None


# ---------------------------------------------------------------------------
# Wykrywanie formatu
# ---------------------------------------------------------------------------

def detect_parser(filename: str, file_bytes: bytes) -> str | None:
    """
    Zwraca identyfikator parsera jeżeli plik zostanie rozpoznany.
    Zwraca None jeśli należy użyć ogólnego parsera heurystycznego.
    """
    lower = filename.lower()

    # --- Pliki PDF ---
    if file_bytes[:4] == b"%PDF" or lower.endswith(".pdf"):
        return _detect_pdf_parser(file_bytes)

    # Goldman Sachs TFI — charakterystyczna nazwa pliku
    if "goldman_sachs" in lower or "goldman sachs" in lower:
        return "goldman_sachs_tfi"

    # SUPERFUND TFI XLSX — charakterystyczna nazwa pliku (musi być przed ogólnym xlsx)
    if "superfund" in lower and (lower.endswith(".xlsx") or lower.endswith(".xls")):
        return "superfund_xlsx"

    # PKO TFI — charakterystyczna nazwa pliku
    if "pko" in lower and (lower.endswith(".xlsx") or lower.endswith(".xls")):
        return "pko_tfi"

    # BNP Paribas TFI — charakterystyczna nazwa pliku
    if "bnp" in lower or "bnp_paribas" in lower:
        return "bnp_paribas_tfi"

    # Generali TFI — charakterystyczna nazwa pliku
    if "generali" in lower:
        return "generali_tfi"

    # Erste TFI — charakterystyczna nazwa pliku
    if "erste" in lower:
        return "erste_tfi"

    # Noble Funds NFO — charakterystyczna nazwa pliku
    if "noble" in lower and (lower.endswith(".xlsx") or lower.endswith(".xls")):
        return "noble_nfo"

    # UNIQA TFI xlsx — charakterystyczna nazwa pliku
    if "uniqa" in lower and (lower.endswith(".xlsx") or lower.endswith(".xls")):
        return "uniqa_tfi_xlsx"

    # Weryfikacja przez nagłówek pliku
    try:
        from app.services.parsers.alfa_sfio import _fix_xlsx_bytes, EXPECTED_HEADER_COL0
        fixed = _fix_xlsx_bytes(file_bytes)
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(fixed), read_only=True, data_only=True)
        if wb.sheetnames:
            ws = wb.worksheets[0]
            row_iter = ws.iter_rows(values_only=True)
            first_row = next(row_iter, None)
            second_row = next(row_iter, None)
            # Noble Funds NFO / UNIQA TFI xlsx: col0='Kod IZFiA', col1='Nazwa funduszu'
            # Noble ma dodatkowo col4='ISIN funduszu (id krajowy)'
            if first_row:
                col0 = str(first_row[0]).strip() if first_row[0] else ""
                col1 = str(first_row[1]).strip() if len(first_row) > 1 and first_row[1] else ""
                col4 = str(first_row[4]).strip() if len(first_row) > 4 and first_row[4] else ""
                if col0 == "Kod IZFiA" and col1 == "Nazwa funduszu":
                    if col4 == "ISIN funduszu (id krajowy)":
                        return "noble_nfo"
                    return "uniqa_tfi_xlsx"

            # Pekao TFI: nazwa arkusza zawiera 'Pekao' (np. 'FI_Pekao_WSZYSTKIE')
            if ws.title and "pekao" in ws.title.lower():
                return "pekao_tfi"

            # Erste TFI: wiersz 1 col0 = 'Skład portfeli na dzień:', arkusz 'Zestawienie'
            if first_row:
                col0 = str(first_row[0]).strip() if first_row[0] else ""
                if col0 == "Skład portfeli na dzień:" and ws.title == "Zestawienie":
                    return "erste_tfi"

            # BNP Paribas TFI: wiersz 1 col5 = 'Data wyceny'
            if first_row and len(first_row) > 5:
                col5 = str(first_row[5]).strip() if first_row[5] else ""
                if col5 == "Data wyceny":
                    return "bnp_paribas_tfi"

            # Generali TFI: wiersz 1 col1 = 'Nazwa funduszu / subfunduszu'
            if first_row and len(first_row) > 1:
                col1 = str(first_row[1]).strip() if first_row[1] else ""
                if col1 == "Nazwa funduszu / subfunduszu":
                    return "generali_tfi"

            # Sprawdzenia po col0 — działają też gdy openpyxl zwróci tylko 1 kolumnę
            # (pliki z dimension ref="A1" zwracają tuple z 1 elementem w read_only)
            if first_row:
                col0 = str(first_row[0]).strip() if first_row[0] else ""
                # PKO TFI / ALFA SFIO: wiersz 1 col0 = 'Identyfikator funduszu lub subfunduszu'
                if col0 == "Identyfikator funduszu lub subfunduszu":
                    return "pko_tfi"
                if col0 == "Nazwa Funduszu / Nazwa Subfunduszu":
                    return "goldman_sachs_tfi"
                if col0 == EXPECTED_HEADER_COL0:
                    return "alfa_sfio"
                if col0 == "Kod IZFiA":
                    return "noble_nfo"
            # Alior SFIO: wiersz 1 to tytuł, wiersz 2 to nagłówki
            if second_row:
                from app.services.parsers.alior_sfio import EXPECTED_HEADER_COL0 as ALIOR_HEADER
                col0_2 = str(second_row[0]).strip() if second_row[0] else ""
                if col0_2 == ALIOR_HEADER:
                    return "alior_sfio"

            # PZU TFI: wiersze 1-4 puste/tytuł, wiersz 5 to nagłówki
            # Czytamy dalej przez iterator
            third_row = next(row_iter, None)
            fourth_row = next(row_iter, None)
            fifth_row = next(row_iter, None)
            if fifth_row:
                from app.services.parsers.pzu_tfi import (
                    EXPECTED_HEADER_COL0 as PZU_HEADER_COL0,
                    EXPECTED_HEADER_COL1 as PZU_HEADER_COL1,
                )
                col0_5 = str(fifth_row[0]).strip() if fifth_row[0] else ""
                col1_5 = str(fifth_row[1]).strip() if len(fifth_row) > 1 and fifth_row[1] else ""
                if col0_5 == PZU_HEADER_COL0 and col1_5 == PZU_HEADER_COL1:
                    return "pzu_tfi"
    except Exception:
        pass

    return None


# ---------------------------------------------------------------------------
# Pobranie listy subfunduszy z pliku (przed zapisem)
# ---------------------------------------------------------------------------

def list_subfunds_from_file(parser_id: str, file_bytes: bytes) -> list[str]:
    """
    Dla parserów multi-funduszowych zwraca listę dostępnych subfunduszy.
    """
    if parser_id == "goldman_sachs_tfi":
        from app.services.parsers.goldman_sachs_tfi import list_subfunds
        return list_subfunds(file_bytes)
    if parser_id == "alfa_sfio":
        from app.services.parsers.alfa_sfio import list_subfunds as alfa_list
        return alfa_list(file_bytes)
    if parser_id == "noble_nfo":
        from app.services.parsers.noble_nfo import list_subfunds as noble_list
        return noble_list(file_bytes)
    if parser_id == "alior_sfio":
        from app.services.parsers.alior_sfio import list_subfunds as alior_list
        return alior_list(file_bytes)
    if parser_id == "pzu_tfi":
        from app.services.parsers.pzu_tfi import list_subfunds as pzu_list
        return pzu_list(file_bytes)
    if parser_id == "uniqa_fio":
        from app.services.parsers.uniqa_fio import list_subfunds as uniqa_list
        return uniqa_list(file_bytes)
    if parser_id == "superfund_tfi":
        from app.services.parsers.superfund_tfi import list_subfunds as superfund_list
        return superfund_list(file_bytes)
    if parser_id == "superfund_xlsx":
        from app.services.parsers.superfund_xlsx import list_subfunds as superfund_xlsx_list
        return superfund_xlsx_list(file_bytes)
    if parser_id == "generali_tfi":
        from app.services.parsers.generali_tfi import list_subfunds as generali_list
        return generali_list(file_bytes)
    if parser_id == "pko_tfi":
        from app.services.parsers.pko_tfi import list_subfunds as pko_list
        return pko_list(file_bytes)
    if parser_id == "bnp_paribas_tfi":
        from app.services.parsers.bnp_paribas_tfi import list_subfunds as bnp_list
        return bnp_list(file_bytes)
    if parser_id == "erste_tfi":
        from app.services.parsers.erste_tfi import list_subfunds as erste_list
        return erste_list(file_bytes)
    if parser_id == "pekao_tfi":
        from app.services.parsers.pekao_tfi import list_subfunds as pekao_list
        return pekao_list(file_bytes)
    if parser_id == "uniqa_tfi_xlsx":
        from app.services.parsers.uniqa_tfi_xlsx import list_subfunds as uniqa_xlsx_list
        return uniqa_xlsx_list(file_bytes)
    return []


# ---------------------------------------------------------------------------
# Parsowanie z użyciem konkretnego parsera
# ---------------------------------------------------------------------------

def parse_with_parser(
    parser_id: str,
    file_bytes: bytes,
    filename: str = "",
    subfund_filter: str | None = None,
) -> list[ParsedPortfolio]:
    """
    Parsuje plik z użyciem zidentyfikowanego parsera.
    Zwraca listę ParsedPortfolio (jeden per subfundusz lub jeden ogólny).
    Jeśli subfund_filter podany → zwraca co najwyżej jeden element.
    """
    if parser_id == "goldman_sachs_tfi":
        from app.services.parsers.goldman_sachs_tfi import parse_goldman_sachs
        snapshots = parse_goldman_sachs(
            file_bytes,
            filename=filename,
            subfund_filter=subfund_filter,
        )
        return [_goldman_to_parsed(s) for s in snapshots]

    if parser_id == "alfa_sfio":
        from app.services.parsers.alfa_sfio import parse_alfa_sfio
        snapshots = parse_alfa_sfio(
            file_bytes,
            filename=filename,
            subfund_filter=subfund_filter,
        )
        return [_alfa_to_parsed(s) for s in snapshots]

    if parser_id == "noble_nfo":
        from app.services.parsers.noble_nfo import parse_noble_nfo
        snapshots = parse_noble_nfo(
            file_bytes,
            filename=filename,
            subfund_filter=subfund_filter,
        )
        return [_noble_to_parsed(s) for s in snapshots]

    if parser_id == "alior_sfio":
        from app.services.parsers.alior_sfio import parse_alior_sfio
        snapshots = parse_alior_sfio(
            file_bytes,
            filename=filename,
            subfund_filter=subfund_filter,
        )
        return [_alior_to_parsed(s) for s in snapshots]

    if parser_id == "pzu_tfi":
        from app.services.parsers.pzu_tfi import parse_pzu_tfi
        snapshots = parse_pzu_tfi(
            file_bytes,
            filename=filename,
            subfund_filter=subfund_filter,
        )
        return [_pzu_to_parsed(s) for s in snapshots]

    if parser_id == "uniqa_fio":
        from app.services.parsers.uniqa_fio import parse_uniqa_fio
        snapshots = parse_uniqa_fio(
            file_bytes,
            filename=filename,
            subfund_filter=subfund_filter,
        )
        return [_uniqa_to_parsed(s) for s in snapshots]

    if parser_id == "superfund_tfi":
        from app.services.parsers.superfund_tfi import parse_superfund_tfi
        snapshots = parse_superfund_tfi(
            file_bytes,
            filename=filename,
            subfund_filter=subfund_filter,
        )
        return [_superfund_to_parsed(s) for s in snapshots]

    if parser_id == "superfund_xlsx":
        from app.services.parsers.superfund_xlsx import parse_superfund_xlsx
        snapshots = parse_superfund_xlsx(
            file_bytes,
            filename=filename,
            subfund_filter=subfund_filter,
        )
        return [_superfund_xlsx_to_parsed(s) for s in snapshots]

    if parser_id == "generali_tfi":
        from app.services.parsers.generali_tfi import parse_generali_tfi
        snapshots = parse_generali_tfi(
            file_bytes,
            subfund_filter=subfund_filter,
        )
        return [_generali_to_parsed(s) for s in snapshots]

    if parser_id == "pko_tfi":
        from app.services.parsers.pko_tfi import parse_pko_tfi
        snapshots = parse_pko_tfi(
            file_bytes,
            subfund_filter=subfund_filter,
        )
        return [_pko_to_parsed(s) for s in snapshots]

    if parser_id == "bnp_paribas_tfi":
        from app.services.parsers.bnp_paribas_tfi import parse_bnp_paribas_tfi
        snapshots = parse_bnp_paribas_tfi(
            file_bytes,
            subfund_filter=subfund_filter,
        )
        return [_bnp_to_parsed(s) for s in snapshots]

    if parser_id == "erste_tfi":
        from app.services.parsers.erste_tfi import parse_erste_tfi
        snapshots = parse_erste_tfi(
            file_bytes,
            subfund_filter=subfund_filter,
        )
        return [_erste_to_parsed(s) for s in snapshots]

    if parser_id == "uniqa_tfi_xlsx":
        from app.services.parsers.uniqa_tfi_xlsx import parse_uniqa_xlsx
        snapshots = parse_uniqa_xlsx(
            file_bytes,
            filename=filename,
            subfund_filter=subfund_filter,
        )
        return [_uniqa_xlsx_to_parsed(s) for s in snapshots]

    if parser_id == "pekao_tfi":
        from app.services.parsers.pekao_tfi import parse_pekao_tfi
        snapshots = parse_pekao_tfi(
            file_bytes,
            subfund_filter=subfund_filter,
        )
        return [_pekao_to_parsed(s) for s in snapshots]

    raise ValueError(f"Nieznany parser: {parser_id}")


# ---------------------------------------------------------------------------
# Czy parser obsługuje wiele subfunduszy?
# ---------------------------------------------------------------------------

MULTI_FUND_PARSERS = {"goldman_sachs_tfi", "alfa_sfio", "noble_nfo", "alior_sfio", "pzu_tfi", "uniqa_fio", "superfund_tfi", "superfund_xlsx", "erste_tfi", "uniqa_tfi_xlsx", "bnp_paribas_tfi", "generali_tfi", "pko_tfi", "pekao_tfi"}


def is_multi_fund(parser_id: str) -> bool:
    return parser_id in MULTI_FUND_PARSERS
