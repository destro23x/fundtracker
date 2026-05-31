"""
Parser dla Generali Fundusze FIO — format „Skład portfela".

Struktura pliku:
  Arkusz: „FIO" (lub inny, ale ze specyficznym nagłówkiem)
  Row 0:  nagłówki kolumn (16 kolumn)
  Row 1+: dane pozycji — wiele subfunduszy w jednym arkuszu

Kolumny (0-based):
  0  - Identyfikator Subfunduszu            → fund_id  (np. „7103")
  1  - Nazwa funduszu / subfunduszu         → subfund_name
  2  - Typ funduszu                         → fund_type (pełna nazwa → normalizacja do FIO/SFIO/…)
  3  - Standardowe identyfikatory subf.     (pominięte, często „N/D")
  4  - Waluta wyceny aktywów                → currency_fund
  5  - Nazwa emitenta                       → company_name
  6  - Identyfikator instrumentu (ISIN)     → isin
  7  - Inny identyfikator                   (pominięty)
  8  - Typ instrumentu                      → asset_type
  9  - Kategoria instrumentu                (pominięta)
  10 - Kraj emitenta                        → country (pełna nazwa po polsku)
  11 - Waluta instrumentu                   → currency_instrument
  12 - Ilość instrumentów w portfelu        → shares
  13 - Wartość instrumentu w walucie wyceny → value
  14 - Procentowy udział w wartości aktywów → weight_pct (ułamek dziesiętny → ×100 = %)
  15 - Informacje uzupełniające             (pominięte)

Uwaga: plik nie zawiera daty wyceny — pobierana z metadanych xlsx (modified)
       i zaokrąglana do końca poprzedniego kwartału.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Optional


# ---------------------------------------------------------------------------
# Modele wewnętrzne
# ---------------------------------------------------------------------------

@dataclass
class GeneraliPosition:
    company_name: str
    isin: Optional[str]
    asset_type: Optional[str]
    shares: Optional[Decimal]
    value: Optional[Decimal]
    weight_pct: Optional[Decimal]   # już w procentach (po ×100)
    currency: str = "PLN"
    country: Optional[str] = None


@dataclass
class GeneraliSubfundSnapshot:
    subfund_name: str
    umbrella_name: str
    snapshot_date: date
    total_value: Optional[Decimal]
    fund_type: Optional[str] = None
    fund_id: Optional[str] = None
    currency_fund: str = "PLN"
    positions: list[GeneraliPosition] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Sygnatura rozpoznawcza
# ---------------------------------------------------------------------------

DETECTION_HEADER_COL1 = "Nazwa funduszu / subfunduszu"
UMBRELLA_NAME = "Generali Fundusze FIO"

# Normalizacja pełnej nazwy typu na skrót
_TYPE_MAP = {
    "fundusz inwestycyjny otwarty": "FIO",
    "specjalistyczny fundusz inwestycyjny otwarty": "SFIO",
    "fundusz inwestycyjny zamknięty": "FIZ",
    "niestandaryzowany fundusz inwestycyjny zamknięty": "FIZAN",
}


def _normalize_fund_type(raw: str | None) -> str | None:
    if not raw:
        return None
    key = raw.strip().lower()
    return _TYPE_MAP.get(key, raw.strip())


def _to_decimal(v) -> Optional[Decimal]:
    if v is None:
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError):
        return None


def _quarter_end(dt: datetime) -> date:
    """Zwraca datę końca kwartału poprzedzającego podaną datę."""
    m = dt.month
    if m <= 3:
        return date(dt.year - 1, 12, 31)
    elif m <= 6:
        return date(dt.year, 3, 31)
    elif m <= 9:
        return date(dt.year, 6, 30)
    else:
        return date(dt.year, 9, 30)


# ---------------------------------------------------------------------------
# Główna funkcja parsowania
# ---------------------------------------------------------------------------

def parse_generali_tfi(
    file_bytes: bytes,
    subfund_filter: Optional[str] = None,
) -> list[GeneraliSubfundSnapshot]:
    """
    Parsuje plik Generali Fundusze FIO i zwraca listę GeneraliSubfundSnapshot.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    if not wb.sheetnames:
        raise ValueError("Plik Generali TFI nie zawiera arkuszy")

    # Ustal datę snapshot z metadanych pliku
    snapshot_date: Optional[date] = None
    props = wb.properties
    if props and props.modified:
        snapshot_date = _quarter_end(props.modified)
    if snapshot_date is None:
        raise ValueError(
            "Nie można określić daty wyceny z pliku Generali TFI. "
            "Plik nie zawiera kolumny z datą, a metadane są niedostępne."
        )

    # Znajdź arkusz z właściwym nagłówkiem
    ws = None
    for sh_name in wb.sheetnames:
        sh = wb[sh_name]
        first_row = next(sh.iter_rows(values_only=True), None)
        if first_row and len(first_row) > 1:
            col1 = str(first_row[1]).strip() if first_row[1] else ""
            if col1 == DETECTION_HEADER_COL1:
                ws = sh
                break

    if ws is None:
        raise ValueError(
            f"Nie znaleziono arkusza z nagłówkiem '{DETECTION_HEADER_COL1}' w pliku Generali TFI"
        )

    rows = list(ws.iter_rows(values_only=True))

    # Zbierz pozycje per subfundusz (pomijamy wiersz 0 = nagłówki)
    subfund_data: dict[str, list] = {}
    subfund_meta: dict[str, tuple] = {}

    for row in rows[1:]:
        if row is None or all(c is None for c in row):
            continue

        subfund_name = str(row[1]).strip() if row[1] else None
        if not subfund_name:
            continue

        if subfund_filter and subfund_name != subfund_filter:
            continue

        if subfund_name not in subfund_meta:
            fund_id = str(row[0]).strip() if row[0] else None
            fund_type = _normalize_fund_type(str(row[2]).strip() if row[2] else None)
            currency_fund = str(row[4]).strip() if row[4] else "PLN"
            subfund_meta[subfund_name] = (fund_id, fund_type, currency_fund)
            subfund_data[subfund_name] = []

        subfund_data[subfund_name].append(row)

    # Konwertuj do GeneraliSubfundSnapshot
    result: list[GeneraliSubfundSnapshot] = []

    for subfund_name, data_rows in subfund_data.items():
        fund_id, fund_type, currency_fund = subfund_meta[subfund_name]

        positions: list[GeneraliPosition] = []

        for row in data_rows:
            company_name = str(row[5]).strip() if row[5] else None
            if not company_name:
                continue

            isin_raw = str(row[6]).strip() if row[6] else None
            isin = isin_raw if isin_raw and isin_raw != "N/D" and len(isin_raw) == 12 else None

            asset_type = str(row[8]).strip() if row[8] else None
            if asset_type == "N/D":
                asset_type = None

            country = str(row[10]).strip() if len(row) > 10 and row[10] else None
            if country == "N/D":
                country = None

            currency = str(row[11]).strip() if len(row) > 11 and row[11] else currency_fund

            shares = _to_decimal(row[12]) if len(row) > 12 else None
            value = _to_decimal(row[13]) if len(row) > 13 else None

            # weight_pct przechowywany jako ułamek dziesiętny (0.0056 = 0.56%)
            weight_raw = _to_decimal(row[14]) if len(row) > 14 else None
            weight_pct = (weight_raw * Decimal("100")).quantize(Decimal("0.0001")) if weight_raw is not None else None

            positions.append(GeneraliPosition(
                company_name=company_name,
                isin=isin,
                asset_type=asset_type,
                shares=shares,
                value=value,
                weight_pct=weight_pct,
                currency=currency or currency_fund,
                country=country,
            ))

        # total_value = suma wartości pozycji
        total_value: Optional[Decimal] = None
        vals = [p.value for p in positions if p.value is not None]
        if vals:
            total_value = sum(vals)  # type: ignore[misc]

        result.append(GeneraliSubfundSnapshot(
            subfund_name=subfund_name,
            umbrella_name=UMBRELLA_NAME,
            snapshot_date=snapshot_date,
            total_value=total_value,
            fund_type=fund_type,
            fund_id=fund_id,
            currency_fund=currency_fund,
            positions=positions,
        ))

    return result


# ---------------------------------------------------------------------------
# Pomocnicze funkcje wymagane przez __init__.py
# ---------------------------------------------------------------------------

def list_subfunds(file_bytes: bytes) -> list[str]:
    """Zwraca listę nazw subfunduszy w pliku bez pełnego parsowania."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    subfunds: list[str] = []
    seen: set[str] = set()

    for sh_name in wb.sheetnames:
        sh = wb[sh_name]
        rows = sh.iter_rows(values_only=True)
        header = next(rows, None)
        if not header or len(header) < 2:
            continue
        col1 = str(header[1]).strip() if header[1] else ""
        if col1 != DETECTION_HEADER_COL1:
            continue
        for row in rows:
            if row and row[1]:
                name = str(row[1]).strip()
                if name and name not in seen:
                    seen.add(name)
                    subfunds.append(name)

    return subfunds
