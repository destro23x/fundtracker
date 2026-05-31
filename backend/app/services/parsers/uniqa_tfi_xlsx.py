"""
Parser dla UNIQA TFI — format xlsx „Portfel UNIQA FIO".

Struktura pliku:
  Row 1:  nagłówki kolumn (16 kolumn)
  Row 2+: dane pozycji

Kolumny (0-based):
  0  - Kod IZFiA
  1  - Nazwa funduszu          → umbrella_name
  2  - Nazwa subfunduszu       → subfund_name
  3  - Typ funduszu
  4  - ISIN funduszu
  5  - Waluta funduszu
  6  - Nazwa emitenta          → company_name
  7  - ISIN instrumentu        → isin
  8  - Inne ID instr.
  9  - Typ instrumentu         → asset_type
  10 - Kategoria instrumentu
  11 - Kraj emitenta
  12 - Waluta instrumentu      → currency
  13 - Ilość                   → shares
  14 - Wartość (PLN)           → value
  15 - Udział w portfelu       → weight_pct

Data snapshot pochodzi z nazwy pliku (format: YYYYMMDD_...) lub — jeśli niedostępna
— musi być podana przez wywołującego.

Detekcja: pierwszy arkusz ma wiersz nagłówkowy gdzie col0='Kod IZFiA' i col1='Nazwa funduszu'.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation
from typing import Optional


# ---------------------------------------------------------------------------
# Modele wewnętrzne
# ---------------------------------------------------------------------------

@dataclass
class UniqaXlsxPosition:
    company_name: str
    isin: Optional[str]
    asset_type: Optional[str]
    shares: Optional[Decimal]
    value: Optional[Decimal]
    weight_pct: Optional[Decimal]
    currency: str = "PLN"


@dataclass
class UniqaXlsxSubfundSnapshot:
    subfund_name: str
    umbrella_name: str
    snapshot_date: Optional[date]
    total_value: Optional[Decimal]
    positions: list[UniqaXlsxPosition] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Sygnatura rozpoznawcza
# ---------------------------------------------------------------------------

EXPECTED_HEADER_COL0 = "Kod IZFiA"
EXPECTED_HEADER_COL1 = "Nazwa funduszu"

_FILENAME_DATE_RE = re.compile(r"(\d{4})(\d{2})(\d{2})")


def _extract_date_from_filename(filename: str) -> Optional[date]:
    m = _FILENAME_DATE_RE.search(filename)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return None


def _to_decimal(v) -> Optional[Decimal]:
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip().rstrip("%").replace(",", ".").replace(" ", "")
        if not v or v == "N/D":
            return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError):
        return None


# ---------------------------------------------------------------------------
# Główna funkcja parsowania
# ---------------------------------------------------------------------------

def parse_uniqa_xlsx(
    file_bytes: bytes,
    filename: str = "",
    subfund_filter: Optional[str] = None,
) -> list[UniqaXlsxSubfundSnapshot]:
    """
    Parsuje plik xlsx UNIQA FIO i zwraca listę UniqaXlsxSubfundSnapshot.
    """
    import openpyxl

    snapshot_date = _extract_date_from_filename(filename)

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if not wb.sheetnames:
        raise ValueError("Plik nie zawiera arkuszy")

    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Plik jest pusty")

    # Pomiń wiersz nagłówkowy (row 0)
    # Zbierz pozycje per subfundusz
    subfund_rows: dict[str, tuple[str, list]] = {}

    for row in rows[1:]:
        if row is None or all(c is None for c in row):
            continue

        subfund_name = str(row[2]).strip() if row[2] and str(row[2]).strip() != "N/D" else None
        umbrella_name = str(row[1]).strip() if row[1] and str(row[1]).strip() != "N/D" else "UNIQA TFI"

        if not subfund_name:
            continue

        if subfund_filter and subfund_name != subfund_filter:
            continue

        if subfund_name not in subfund_rows:
            subfund_rows[subfund_name] = (umbrella_name, [])
        subfund_rows[subfund_name][1].append(row)

    result: list[UniqaXlsxSubfundSnapshot] = []

    for subfund_name, (umbrella_name, data_rows) in subfund_rows.items():
        positions: list[UniqaXlsxPosition] = []

        for row in data_rows:
            company_name = str(row[6]).strip() if row[6] and str(row[6]).strip() not in ("N/D", "") else None
            if not company_name:
                continue

            isin_raw = str(row[7]).strip() if row[7] else None
            isin = isin_raw if isin_raw and len(isin_raw) == 12 and isin_raw != "N/D" else None

            asset_type = str(row[9]).strip() if row[9] and str(row[9]).strip() != "N/D" else None
            currency = str(row[12]).strip() if row[12] and str(row[12]).strip() != "N/D" else "PLN"

            shares = _to_decimal(row[13])
            value = _to_decimal(row[14])

            # Udział w portfelu — może być jako % lub ułamek
            weight_raw = _to_decimal(row[15])
            if weight_raw is not None and weight_raw <= Decimal("1"):
                weight_pct = weight_raw * Decimal("100")
            else:
                weight_pct = weight_raw

            positions.append(UniqaXlsxPosition(
                company_name=company_name,
                isin=isin,
                asset_type=asset_type,
                shares=shares,
                value=value,
                weight_pct=weight_pct.quantize(Decimal("0.0001")) if weight_pct is not None else None,
                currency=currency or "PLN",
            ))

        # Wylicz total_value jako sumę wartości
        total_value: Optional[Decimal] = None
        vals = [p.value for p in positions if p.value is not None]
        if vals:
            total_value = sum(vals)  # type: ignore[misc]

        result.append(UniqaXlsxSubfundSnapshot(
            subfund_name=subfund_name,
            umbrella_name=umbrella_name,
            snapshot_date=snapshot_date,
            total_value=total_value,
            positions=positions,
        ))

    return result


# ---------------------------------------------------------------------------
# Pomocnicze funkcje wymagane przez __init__.py
# ---------------------------------------------------------------------------

def list_subfunds(file_bytes: bytes) -> list[str]:
    """Zwraca listę nazw subfunduszy bez pełnego parsowania."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if not wb.sheetnames:
        return []

    ws = wb.worksheets[0]
    seen: list[str] = []
    seen_set: set[str] = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row is None:
            continue
        name = str(row[2]).strip() if row[2] and str(row[2]).strip() != "N/D" else None
        if name and name not in seen_set:
            seen.append(name)
            seen_set.add(name)

    return seen
