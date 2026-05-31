"""
Parser dla TFI PZU – wielozakładkowy plik KNF z 52+ subfunduszami.

Struktura pliku:
  - Wiele zakładek (jedna na miesiąc, np. "31.03.2026", "31.12.2025" …)
  - Wiersze 1-4: puste lub tytuł
  - Wiersz 5:    nagłówki kolumn
  - Dane od wiersza 6

Mapowanie kolumn (wiersz 5):
  col0   Identyfikator IZFIA funduszu lub subfunduszu
  col1   Nazwa funduszu
  col2   Nazwa subfunduszu        → subfund_name
  col5   Waluta wyceny funduszu
  col6   Emitent                  → company_name
  col7   Kod ISIN instrumentu     → isin
  col9   Typ instrumentu          → asset_type
  col11  Kraj emitenta            → country
  col12  Waluta wyceny instrumentu → currency
  col13  Ilość instrumentów       → shares
  col14  Wartość instrumentu      → value
  col15  Udział procentowy w aktywach (w %)  → weight_pct (ułamek dziesiętny → ×100)
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation

import openpyxl


EXPECTED_HEADER_COL0 = "Identyfikator IZFIA funduszu lub subfunduszu"
EXPECTED_HEADER_COL1 = "Nazwa funduszu"
HEADER_ROW = 5  # 1-indexed


# ---------------------------------------------------------------------------
# Typy danych
# ---------------------------------------------------------------------------

@dataclass
class PzuPosition:
    company_name: str
    isin: str | None
    asset_type: str | None
    country: str | None
    currency: str
    shares: Decimal | None
    value: Decimal | None
    weight_pct: Decimal | None  # w procentach (już po ×100)


@dataclass
class PzuSubfundSnapshot:
    subfund_name: str
    fund_name: str
    fund_id: str | None
    izfia_code: str | None          # Kod IZFiA (kol. 0), np. PZU001
    fund_type: str | None
    currency_fund: str | None
    snapshot_date: date
    total_value: Decimal | None
    positions: list[PzuPosition] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Pomocnicze
# ---------------------------------------------------------------------------

def _to_decimal(val) -> Decimal | None:
    if val is None:
        return None
    s = str(val).strip().replace(",", ".").replace("\xa0", "").replace(" ", "")
    if not s or s.lower() in ("n/d", "n/a", "-", ""):
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _clean_str(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ("n/d", "n/a", "") else None


def _extract_date_from_sheet_name(sheet_name: str) -> date | None:
    """Parsuje datę z nazwy zakładki: '31.03.2026' lub '31.01.2025 r.'"""
    m = re.search(r"(\d{1,2})\.(\d{2})\.(\d{4})", sheet_name)
    if m:
        try:
            return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
        except ValueError:
            pass
    return None


# ---------------------------------------------------------------------------
# Główne funkcje
# ---------------------------------------------------------------------------

def list_subfunds(file_bytes: bytes) -> list[str]:
    """Zwraca listę nazw subfunduszy z najnowszej (pierwszej) zakładki."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[0]
    subfunds: list[str] = []
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        name = _clean_str(row[2]) if len(row) > 2 else None
        if name and name not in seen:
            seen.add(name)
            subfunds.append(name)
    return subfunds


def parse_pzu_tfi(
    file_bytes: bytes,
    filename: str = "",
    subfund_filter: str | None = None,
) -> list[PzuSubfundSnapshot]:
    """
    Parsuje plik PZU TFI. Używa tylko pierwszej (najnowszej) zakładki.
    Jeśli subfund_filter podany – zwraca tylko ten subfundusz.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.worksheets[0]

    snapshot_date = _extract_date_from_sheet_name(ws.title)
    if snapshot_date is None:
        m = re.search(r"(\d{4})[-_.](\d{2})[-_.](\d{2})", filename)
        snapshot_date = (
            date(int(m.group(1)), int(m.group(2)), int(m.group(3))) if m else date.today()
        )

    subfunds: dict[str, PzuSubfundSnapshot] = {}

    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if not row or (row[0] is None and row[6] is None):
            continue

        subfund_name = _clean_str(row[2]) if len(row) > 2 else None
        if not subfund_name:
            continue
        if subfund_filter and subfund_name != subfund_filter:
            continue

        company_name = _clean_str(row[6]) if len(row) > 6 else None
        if not company_name:
            continue

        isin_raw = _clean_str(row[7]) if len(row) > 7 else None
        isin = isin_raw if isin_raw and re.match(r"^[A-Z]{2}[A-Z0-9]{10}$", isin_raw) else None

        asset_type = _clean_str(row[9]) if len(row) > 9 else None
        currency = _clean_str(row[12]) if len(row) > 12 else "PLN"
        shares = _to_decimal(row[13]) if len(row) > 13 else None
        value = _to_decimal(row[14]) if len(row) > 14 else None
        weight_raw = _to_decimal(row[15]) if len(row) > 15 else None
        weight_pct = weight_raw * Decimal("100") if weight_raw is not None else None

        if subfund_name not in subfunds:
            subfunds[subfund_name] = PzuSubfundSnapshot(
                subfund_name=subfund_name,
                fund_name=_clean_str(row[1]) or "PZU TFI",
                fund_id=_clean_str(row[4]) if len(row) > 4 else None,
                izfia_code=_clean_str(row[0]) if len(row) > 0 else None,
                fund_type=_clean_str(row[3]) if len(row) > 3 else None,
                currency_fund=_clean_str(row[5]) if len(row) > 5 else "PLN",
                snapshot_date=snapshot_date,
                total_value=None,
            )

        subfunds[subfund_name].positions.append(
            PzuPosition(
                company_name=company_name,
                isin=isin,
                asset_type=asset_type,
                country=_clean_str(row[11]) if len(row) > 11 else None,
                currency=currency or "PLN",
                shares=shares,
                value=value,
                weight_pct=weight_pct,
            )
        )

    return list(subfunds.values())
