"""
Parser dla Erste TFI — format „Zestawienie składu portfeli".

Struktura pliku:
  Row 1:  'Skład portfeli na dzień:' | None | <datetime>  → data snapshot
  Row 2:  nagłówki kolumn (16 kolumn)
  Row 3+: dane pozycji — wiele subfunduszy w jednym arkuszu

Kolumny (0-based):
  0  - Identyfikator funduszu lub Subfunduszu (zazwyczaj puste)
  1  - Pełna nazwa funduszu
  2  - Nazwa subfunduszu
  3  - Typ funduszu
  4  - Standardowy identyfikator subfunduszu
  5  - Waluta wyceny aktywów
  6  - Nazwa emitenta          → company_name
  7  - Kod ISIN instrumentu    → isin
  8  - Inny identyfikator
  9  - Typ instrumentu         → asset_type
  10 - Kategoria instrumentu
  11 - Kraj emitenta
  12 - Waluta wyceny instrumentu → currency
  13 - Ilość instrumentu        → shares
  14 - Wartość instrumentu      → value
  15 - Informacje uzupełniające

Uwaga: brak kolumny weight_pct — jest obliczana jako value/total_value*100
       dla każdego subfunduszu osobno.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Optional


# ---------------------------------------------------------------------------
# Modele wewnętrzne
# ---------------------------------------------------------------------------

@dataclass
class ErstePosition:
    company_name: str
    isin: Optional[str]
    asset_type: Optional[str]
    shares: Optional[Decimal]
    value: Optional[Decimal]
    weight_pct: Optional[Decimal]
    currency: str = "PLN"
    country: Optional[str] = None


@dataclass
class ErsteSubfundSnapshot:
    subfund_name: str
    umbrella_name: str
    snapshot_date: date
    total_value: Optional[Decimal]
    fund_type: Optional[str] = None
    fund_id: Optional[str] = None
    positions: list[ErstePosition] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Sygnatura rozpoznawcza
# ---------------------------------------------------------------------------

EXPECTED_SHEET = "Zestawienie"
EXPECTED_DATE_LABEL = "Skład portfeli na dzień:"
EXPECTED_HEADER_COL6 = "Nazwa emitenta"


def _to_decimal(v) -> Optional[Decimal]:
    if v is None:
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError):
        return None


# ---------------------------------------------------------------------------
# Główna funkcja parsowania
# ---------------------------------------------------------------------------

def parse_erste_tfi(
    file_bytes: bytes,
    subfund_filter: Optional[str] = None,
) -> list[ErsteSubfundSnapshot]:
    """
    Parsuje plik Erste TFI i zwraca listę ErsteSubfundSnapshot
    (jeden per subfundusz, opcjonalnie filtrowany po nazwie).
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    if EXPECTED_SHEET not in wb.sheetnames:
        raise ValueError(f"Brak arkusza '{EXPECTED_SHEET}' w pliku Erste TFI")

    ws = wb[EXPECTED_SHEET]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Plik jest pusty")

    # Wiersz 1: data snapshot
    snapshot_date: Optional[date] = None
    first_row = rows[0]
    if first_row and len(first_row) > 2 and first_row[2] is not None:
        raw_date = first_row[2]
        if isinstance(raw_date, datetime):
            snapshot_date = raw_date.date()
        elif isinstance(raw_date, date):
            snapshot_date = raw_date

    if snapshot_date is None:
        raise ValueError("Nie można odczytać daty snapshot z pliku Erste TFI")

    # Zbierz pozycje per subfundusz
    # key: subfund_name → (umbrella_name, [raw_rows])
    subfund_rows: dict[str, tuple[str, list]] = {}

    for row in rows[2:]:  # pomijamy row1 (data) i row2 (nagłówki)
        if row is None or all(c is None for c in row):
            continue

        subfund_name = str(row[2]).strip() if row[2] else None
        umbrella_name = str(row[1]).strip() if row[1] else "Erste TFI"
        fund_type = str(row[3]).strip() if len(row) > 3 and row[3] else None
        fund_id = str(row[4]).strip() if len(row) > 4 and row[4] else None

        if not subfund_name:
            continue

        if subfund_filter and subfund_name != subfund_filter:
            continue

        if subfund_name not in subfund_rows:
            subfund_rows[subfund_name] = (umbrella_name, fund_type, fund_id, [])
        subfund_rows[subfund_name][3].append(row)

    # Konwertuj do ErsteSubfundSnapshot
    result: list[ErsteSubfundSnapshot] = []

    for subfund_name, (umbrella_name, fund_type, fund_id, data_rows) in subfund_rows.items():
        positions: list[ErstePosition] = []

        for row in data_rows:
            company_name = str(row[6]).strip() if row[6] else None
            if not company_name:
                continue

            isin_raw = str(row[7]).strip() if row[7] else None
            isin = isin_raw if isin_raw and len(isin_raw) == 12 else None

            asset_type_raw = str(row[9]).strip() if row[9] else None
            country = str(row[11]).strip() if len(row) > 11 and row[11] else None
            currency = str(row[12]).strip() if row[12] else "PLN"

            shares = _to_decimal(row[13])
            value = _to_decimal(row[14])

            positions.append(ErstePosition(
                company_name=company_name,
                isin=isin,
                asset_type=asset_type_raw,
                shares=shares,
                value=value,
                weight_pct=None,  # obliczone poniżej
                currency=currency or "PLN",
                country=country,
            ))

        # Oblicz total_value i weight_pct
        total_value: Optional[Decimal] = None
        values_with_val = [p for p in positions if p.value is not None]
        if values_with_val:
            total_value = sum(p.value for p in values_with_val)  # type: ignore[misc]
            if total_value and total_value > 0:
                for p in positions:
                    if p.value is not None:
                        p.weight_pct = (p.value / total_value * Decimal("100")).quantize(Decimal("0.0001"))

        result.append(ErsteSubfundSnapshot(
            subfund_name=subfund_name,
            umbrella_name=umbrella_name,
            snapshot_date=snapshot_date,
            total_value=total_value,
            fund_type=fund_type,
            fund_id=fund_id,
            positions=positions,
        ))

    return result


# ---------------------------------------------------------------------------
# Pomocnicze funkcje wymagane przez __init__.py
# ---------------------------------------------------------------------------

def list_subfunds(file_bytes: bytes) -> list[str]:
    """Zwraca listę nazw subfunduszy w pliku bez pełnego parsowania."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    if EXPECTED_SHEET not in wb.sheetnames:
        return []

    ws = wb[EXPECTED_SHEET]
    seen: list[str] = []
    seen_set: set[str] = set()

    for row in ws.iter_rows(min_row=3, values_only=True):
        if row is None:
            continue
        name = str(row[2]).strip() if row[2] else None
        if name and name not in seen_set:
            seen.append(name)
            seen_set.add(name)

    return seen
