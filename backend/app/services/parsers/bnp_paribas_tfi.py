"""
Parser dla BNP Paribas TFI — format „Składy portfeli funduszy".

Struktura pliku:
  Jeden arkusz (nazwa zmienna, np. „Składy potfeli 1Q 2026").
  Row 0:  nagłówki kolumn (18 kolumn)
  Row 1+: dane pozycji — wiele subfunduszy w jednym arkuszu

Kolumny (0-based):
  0  - Identyfikator funduszu lub Subfunduszu → izfia_code (np. "PLFIO000001")
  1  - Pełna nazwa funduszu                   → umbrella_name
  2  - Nazwa subfunduszu                       → subfund_name
  3  - Typ funduszu                            → fund_type
  4  - Standardowy identyfikator subfunduszu   (często pusty)
  5  - Data wyceny                             → snapshot_date (per row)
  6  - Waluta wyceny aktywów i zobowiązań      → currency_fund
  7  - Nazwa emitenta                          → company_name
  8  - Identyfikator instrumentu (ISIN)        → isin
  9  - Inny identyfikator                      (pominięty)
  10 - Typ instrumentu                         → asset_type
  11 - Kategoria instrumentu                   (pominięta)
  12 - Kraj emitenta                           → country
  13 - Waluta wyceny instrumentu               → currency_instrument
  14 - Ilość instrumentu w portfelu            → shares
  15 - Wartość instrumentu w walucie funduszu  → value
  16 - Procentowy udział w Aktywach ogółem     → weight_pct (%)
  17 - Informacje uzupełniające                (pominięte)
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Optional


# ---------------------------------------------------------------------------
# Modele wewnętrzne
# ---------------------------------------------------------------------------

@dataclass
class BnpPosition:
    company_name: str
    isin: Optional[str]
    asset_type: Optional[str]
    shares: Optional[Decimal]
    value: Optional[Decimal]
    weight_pct: Optional[Decimal]
    currency: str = "PLN"
    country: Optional[str] = None


@dataclass
class BnpSubfundSnapshot:
    subfund_name: str
    umbrella_name: str
    snapshot_date: date
    total_value: Optional[Decimal]
    fund_type: Optional[str] = None
    fund_id: Optional[str] = None        # zawsze None — BNP używa izfia_code
    izfia_code: Optional[str] = None     # col 0: np. "PLFIO000001" (identyfikator IZFiA)
    currency_fund: str = "PLN"
    positions: list[BnpPosition] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Sygnatura rozpoznawcza (używana w detect_parser)
# ---------------------------------------------------------------------------

DETECTION_HEADER_COL5 = "Data wyceny"
DETECTION_HEADER_COL1 = "Pełna"   # fragment "Pełna \nnazwa \nfunduszu"


def _to_decimal(v) -> Optional[Decimal]:
    if v is None:
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError):
        return None


def _to_date(v) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


# ---------------------------------------------------------------------------
# Główna funkcja parsowania
# ---------------------------------------------------------------------------

def parse_bnp_paribas_tfi(
    file_bytes: bytes,
    subfund_filter: Optional[str] = None,
) -> list[BnpSubfundSnapshot]:
    """
    Parsuje plik BNP Paribas TFI i zwraca listę BnpSubfundSnapshot
    (jeden per subfundusz, opcjonalnie filtrowany po nazwie).
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)

    if not wb.sheetnames:
        raise ValueError("Plik BNP Paribas TFI nie zawiera arkuszy")

    # Znajdź odpowiedni arkusz — ten, który ma "Data wyceny" w wierszu nagłówkowym
    ws = None
    for sh_name in wb.sheetnames:
        sh = wb[sh_name]
        first_row = next(sh.iter_rows(values_only=True), None)
        if first_row and len(first_row) > 5:
            col5 = str(first_row[5]).strip() if first_row[5] else ""
            if col5 == DETECTION_HEADER_COL5:
                ws = sh
                break

    if ws is None:
        raise ValueError("Nie znaleziono arkusza z kolumną 'Data wyceny' w pliku BNP Paribas TFI")

    rows = list(ws.iter_rows(values_only=True))

    # Zbierz pozycje per subfundusz (pomijamy wiersz 0 = nagłówki)
    # key: subfund_name → (umbrella_name, fund_id, fund_type, currency_fund, snapshot_date, [rows])
    subfund_data: dict[str, list] = {}
    subfund_meta: dict[str, tuple] = {}

    for row in rows[1:]:
        if row is None or all(c is None for c in row):
            continue

        subfund_name = str(row[2]).strip() if row[2] else None
        if not subfund_name:
            continue

        if subfund_filter and subfund_name != subfund_filter:
            continue

        if subfund_name not in subfund_meta:
            umbrella_name = str(row[1]).strip() if row[1] else "BNP Paribas TFI"
            izfia_code = str(row[0]).strip() if row[0] else None  # col 0: "PLFIO000001" = kod IZFiA
            fund_type = str(row[3]).strip() if row[3] else None
            currency_fund = str(row[6]).strip() if row[6] else "PLN"
            snapshot_date = _to_date(row[5])
            subfund_meta[subfund_name] = (umbrella_name, izfia_code, fund_type, currency_fund, snapshot_date)
            subfund_data[subfund_name] = []

        subfund_data[subfund_name].append(row)

    # Konwertuj do BnpSubfundSnapshot
    result: list[BnpSubfundSnapshot] = []

    for subfund_name, data_rows in subfund_data.items():
        umbrella_name, izfia_code, fund_type, currency_fund, snapshot_date = subfund_meta[subfund_name]

        if snapshot_date is None:
            raise ValueError(f"Brak daty wyceny dla subfunduszu '{subfund_name}'")

        positions: list[BnpPosition] = []

        for row in data_rows:
            company_name = str(row[7]).strip() if row[7] else None
            if not company_name:
                continue

            isin_raw = str(row[8]).strip() if row[8] else None
            isin = isin_raw if isin_raw and len(isin_raw) == 12 else None

            asset_type = str(row[10]).strip() if row[10] else None
            country = str(row[12]).strip() if len(row) > 12 and row[12] else None
            currency = str(row[13]).strip() if len(row) > 13 and row[13] else currency_fund

            shares = _to_decimal(row[14]) if len(row) > 14 else None
            value = _to_decimal(row[15]) if len(row) > 15 else None
            weight_pct = _to_decimal(row[16]) if len(row) > 16 else None

            positions.append(BnpPosition(
                company_name=company_name,
                isin=isin,
                asset_type=asset_type,
                shares=shares,
                value=value,
                weight_pct=weight_pct,
                currency=currency or currency_fund,
                country=country,
            ))

        # total_value = suma wartości pozycji
        total_value: Optional[Decimal] = None
        vals = [p.value for p in positions if p.value is not None]
        if vals:
            total_value = sum(vals)  # type: ignore[misc]

        result.append(BnpSubfundSnapshot(
            subfund_name=subfund_name,
            umbrella_name=umbrella_name,
            snapshot_date=snapshot_date,
            total_value=total_value,
            fund_type=fund_type,
            fund_id=None,
            izfia_code=izfia_code,
            currency_fund=currency_fund,
            positions=positions,
        ))

    return result


# ---------------------------------------------------------------------------
# Pomocnicze funkcje wymagane przez __init__.py
# ---------------------------------------------------------------------------

def list_subfunds(file_bytes: bytes) -> list[str]:
    """Zwraca listę nazw subfunduszy w pliku bez pełnego parsowania."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    subfunds: list[str] = []
    seen: set[str] = set()

    for sh_name in wb.sheetnames:
        sh = wb[sh_name]
        rows = sh.iter_rows(values_only=True)
        header = next(rows, None)
        if not header or len(header) <= 5:
            continue
        col5 = str(header[5]).strip() if header[5] else ""
        if col5 != DETECTION_HEADER_COL5:
            continue
        for row in rows:
            if row and row[2]:
                name = str(row[2]).strip()
                if name and name not in seen:
                    seen.add(name)
                    subfunds.append(name)

    return subfunds
