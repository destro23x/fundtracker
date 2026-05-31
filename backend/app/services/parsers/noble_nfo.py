"""
Parser dla pliku Noble Funds TFI / VeloFunds (format KNF regulacyjny NFO).

Format pliku (jeden arkusz 'Portfel'):
  Wiersz 1: nagłówek (16 kolumn)
  Wiersze 2+: dane — jeden wiersz = jedna pozycja

Kolumny:
  [0]  Kod IZFiA                          — ID TFI (np. NOB003)
  [1]  Nazwa funduszu                     — VeloFunds FIO / VeloFunds SFIO
  [2]  Nazwa subfunduszu                  — VeloFund Akcji Polskich …
  [3]  Typ funduszu                       — FIO / SFIO
  [4]  ISIN funduszu (id krajowy)         — ISIN parasola
  [5]  Waluta wyceny fund.                — PLN
  [6]  Nazwa emitenta                     — nazwa spółki
  [7]  ISIN instrumentu                   — ISIN akcji/obligacji (lub N/D)
  [8]  Inne ID instr.                     — N/D lub alternatywny ID
  [9]  Typ instrumentu                    — Akcje, Obligacje, Instrumenty pochodne…
  [10] Kategoria instrumentu
  [11] Kraj emitenta                      — POLSKA, LUKSEMBURG… lub N/D
  [12] Waluta wyceny instr.               — PLN, EUR, SEK…
  [13] Ilość instr.                       — ilość (może być ujemna dla pochodnych)
  [14] Wartość instr.                     — wartość w walucie wyceny funduszu
  [15] Informacje uzupełniające

Uwaga: plik nie zawiera kolumny z udziałem procentowym.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation

import openpyxl

# ---------------------------------------------------------------------------
# Stałe kolumn
# ---------------------------------------------------------------------------

COL_IZFiA_ID = 0
COL_FUND_NAME = 1
COL_SUBFUND_NAME = 2
COL_FUND_TYPE = 3
COL_FUND_ID = 4   # ISIN funduszu (id krajowy) — np. PLFIO000012
COL_CURRENCY_FUND = 5
COL_COMPANY_NAME = 6
COL_ISIN = 7
COL_ALT_ID = 8
COL_ASSET_TYPE = 9
COL_COUNTRY = 11
COL_CURRENCY_INSTR = 12
COL_SHARES = 13
COL_VALUE = 14

EXPECTED_HEADER_COL0 = "Kod IZFiA"

ND_VALUES = {"n/d", "nd", "n.d.", "-", ""}

# ---------------------------------------------------------------------------
# Typy danych
# ---------------------------------------------------------------------------


@dataclass
class NoblePosition:
    company_name: str
    isin: str | None
    currency: str
    shares: Decimal | None
    value: Decimal | None
    asset_type: str | None
    country: str | None = None


@dataclass
class NobleSubfundSnapshot:
    subfund_name: str
    fund_name: str | None
    izfia_id: str | None
    izfia_code: str | None          # Kod IZFiA (kol. 0), np. NOB003
    fund_type: str | None
    snapshot_date: date | None
    positions: list[NoblePosition] = field(default_factory=list)

    @property
    def total_value(self) -> Decimal:
        return sum(
            (p.value for p in self.positions if p.value is not None),
            Decimal("0"),
        )


# ---------------------------------------------------------------------------
# Pomocnicze
# ---------------------------------------------------------------------------


def _to_decimal(val) -> Decimal | None:
    if val is None:
        return None
    try:
        return Decimal(str(val))
    except InvalidOperation:
        return None


def _nd(val) -> str | None:
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in ND_VALUES:
        return None
    return s


def _normalize_isin(raw: str | None) -> str | None:
    if not raw:
        return None
    s = raw.strip()
    if s.lower() in ND_VALUES:
        return None
    if re.match(r"^[A-Z]{2}[A-Z0-9]{10}$", s):
        return s
    # Lokalny identyfikator — zachowaj jeśli wygląda sensownie
    if len(s) >= 6:
        return s
    return None


def _extract_date_from_filename(filename: str) -> date | None:
    m = re.search(r"(\d{4})-(\d{2})-(\d{2})", filename)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    m = re.search(r"(\d{4})(\d{2})(\d{2})", filename)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass
    return None


# ---------------------------------------------------------------------------
# Główna funkcja parsująca
# ---------------------------------------------------------------------------


def parse_noble_nfo(
    file_bytes: bytes,
    filename: str = "",
    subfund_filter: str | None = None,
) -> list[NobleSubfundSnapshot]:
    """
    Parsuje plik Noble Funds TFI / VeloFunds (format KNF NFO).

    Args:
        file_bytes:     Zawartość pliku .xlsx
        filename:       Nazwa pliku (do wykrycia daty)
        subfund_filter: Jeśli podany — zwraca tylko dopasowany subfundusz

    Returns:
        Lista NobleSubfundSnapshot (po jednym na subfundusz).
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    if not wb.sheetnames:
        raise ValueError("Plik jest pusty (brak arkuszy).")

    # Szukaj arkusza "Portfel" lub weź pierwszy
    sheet_name = "Portfel" if "Portfel" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Arkusz jest pusty.")

    header_val = str(rows[0][COL_IZFiA_ID]).strip() if rows[0][COL_IZFiA_ID] else ""
    if header_val != EXPECTED_HEADER_COL0:
        raise ValueError(
            f"Nieoczekiwany nagłówek (kol. 0): '{header_val}'. "
            f"Oczekiwano: '{EXPECTED_HEADER_COL0}'"
        )

    snapshot_date = _extract_date_from_filename(filename)

    subfunds: dict[str, NobleSubfundSnapshot] = {}

    for row in rows[1:]:
        if not row or row[COL_SUBFUND_NAME] is None:
            continue

        subfund_name = str(row[COL_SUBFUND_NAME]).strip()
        if not subfund_name:
            continue

        if subfund_filter and subfund_filter.lower() not in subfund_name.lower():
            continue

        if subfund_name not in subfunds:
            fund_name = _nd(row[COL_FUND_NAME])
            izfia_id = _nd(row[COL_FUND_ID]) if len(row) > COL_FUND_ID else _nd(row[COL_IZFiA_ID])
            fund_type = _nd(row[COL_FUND_TYPE])
            subfunds[subfund_name] = NobleSubfundSnapshot(
                subfund_name=subfund_name,
                fund_name=fund_name,
                izfia_id=izfia_id,
                izfia_code=_nd(row[COL_IZFiA_ID]),
                fund_type=fund_type,
                snapshot_date=snapshot_date,
            )

        company_name = _nd(row[COL_COMPANY_NAME])
        isin_raw = _nd(row[COL_ISIN])
        # Fallback: sprawdź alt ID jeśli ISIN to N/D
        if not isin_raw:
            isin_raw = _nd(row[COL_ALT_ID])
        asset_type = _nd(row[COL_ASSET_TYPE])
        currency = _nd(row[COL_CURRENCY_INSTR]) or _nd(row[COL_CURRENCY_FUND]) or "PLN"

        shares = _to_decimal(row[COL_SHARES])
        value = _to_decimal(row[COL_VALUE])

        display_name = (
            company_name
            or _normalize_isin(isin_raw)
            or asset_type
            or "Nieznany instrument"
        )

        pos = NoblePosition(
            company_name=display_name[:500],
            isin=_normalize_isin(isin_raw),
            currency=currency,
            shares=shares,
            value=value,
            asset_type=asset_type[:50] if asset_type else None,
            country=_nd(row[COL_COUNTRY]) if len(row) > COL_COUNTRY else None,
        )
        subfunds[subfund_name].positions.append(pos)

    return list(subfunds.values())


def list_subfunds(file_bytes: bytes) -> list[str]:
    """Zwraca listę nazw subfunduszy bez pełnego parsowania."""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if not wb.sheetnames:
        return []
    sheet_name = "Portfel" if "Portfel" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    names: list[str] = []
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[COL_SUBFUND_NAME] is None:
            continue
        name = str(row[COL_SUBFUND_NAME]).strip()
        if name and name not in seen:
            names.append(name)
            seen.add(name)
    return names
